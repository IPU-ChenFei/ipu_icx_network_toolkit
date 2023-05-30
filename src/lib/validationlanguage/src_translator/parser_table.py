import openpyxl
from src.lib.validationlanguage.src_translator.const import *
from src.lib.validationlanguage.src_translator.trans_utils import get_platform_config_real_path, get_bios_dump_file_path
import os
import copy

def to_plain_text(input):
    input = input.strip()
    intab = u'\xa0'
    outab = u' '

    trans = str.maketrans(intab, outab)
    output = input.translate(trans)
    output = output.replace(u'\u200b', '')
    return output.strip()

def get_args(text):
    if len(text.strip()) == 0:
        return []
    args = []
    temp = text.split(',')
    for arg in temp:
        args.append(arg.strip())
    return args



class MappingTableParser(object):
    TAB_NAME_H2L_MAPPING = 'H2L'
    TAB_NAME_L2PY_MAPPING = 'L2PY'
    TAB_NAME_FEATURE = 'Features'
    TAB_NAME_BIOS_MENU = 'BiosMenu'

    def __init__(self, name):
        transfile = get_platform_config_real_path(name)
        print(f'translator mapping table: {transfile}')

        self.filename = name
        self.__file_path = transfile
        self.__wb = None

        self.__h2l_table = None
        self.__l2p_table = None
        self.__feature_table = None
        self.__bios_table = None

    @property
    def workbook(self):
        if self.__wb is None:
            self.__wb = openpyxl.load_workbook(self.__file_path)

        return self.__wb

    @property
    def l2py_table(self):
        if self.__l2p_table is None:
            self.__l2p_table = copy.deepcopy(self.__load_mapping_table(self.TAB_NAME_L2PY_MAPPING))

        Assert(self.__l2p_table is not None, f'Failed to load vl mapping table from tab {self.TAB_NAME_L2PY_MAPPING}')
        return self.__l2p_table

    @property
    def h2l_table(self):
        if self.__h2l_table is None:
            self.__h2l_table = copy.deepcopy(self.__load_mapping_table(self.TAB_NAME_H2L_MAPPING))

        Assert(self.__h2l_table is not None, f'Failed to load high-> low level steps mapping table from tab {self.TAB_NAME_H2L_MAPPING}')
        return self.__h2l_table

    @property
    def high_level_step_os_list(self):
        l = []
        l.extend(self.h2l_table.keys())
        l.append(OP_TCDB)
        return l

    @property
    def low_level_step_os_list(self):
        l = []
        l.extend(self.l2py_table.keys())
        l.append(OP_REPEAT)
        l.append(OP_END)
        return l

    @property
    def feature_table(self):
        if self.__feature_table is None:
            self.load_feature_table()

        Assert(self.__feature_table is not None, f'Failed to load feature definition table from tab {self.TAB_NAME_FEATURE}')
        return self.__feature_table

    @property
    def bios_menu_table(self):
        if self.__bios_table is None:
            self.load_bios_menu_table()

        Assert(self.__bios_table is not None, f'Failed to load bios menu knob definition table from tab {self.TAB_NAME_BIOS_MENU}')
        return self.__bios_table

    def __get_sheet(self, tab_name):
        sheet = None
        for s in self.workbook.worksheets:
            if s.title == tab_name:
                sheet = s
                break

        Assert(sheet is not None, f'Failed to load table from tab {tab_name}')
        return sheet

    def __load_mapping_table(self, tab):
        vl_table = {}
        sheet = self.__get_sheet(tab)
        for line in range(2, sheet.max_row + 1):
            high = to_plain_text(sheet.cell(line, 1).value)
            op, args_str = high.split(':')
            op = op.strip()
            if op not in vl_table.keys():
                vl_table[op] = []
            args = get_args(args_str)
            temp = sheet.cell(line, 2).value
            value = to_plain_text(temp)
            # value = to_plain_text(sheet.cell(line, 2).value)

            low_lines = []
            for l in value.split('\n'):
                if len(l.strip()) > 0:
                    low_lines.append(l.strip())

            vl_table[op].append((args, low_lines))

        return vl_table

    def load_feature_table(self):
        sheet = self.__get_sheet(self.TAB_NAME_FEATURE)

        temp = {}
        for line in range(2, sheet.max_row + 1):
            name = to_plain_text(f'{sheet.cell(line, 1).value}'.strip())
            value = to_plain_text(f'{sheet.cell(line, 2).value}'.strip())
            knob_name = to_plain_text(f'{sheet.cell(line, 3).value}'.strip())
            knob_value = to_plain_text(f'{sheet.cell(line, 4).value}'.strip())

            if (name, value) not in temp.keys():
                temp[(name, value)] = []

            temp[(name, value)].append((knob_name, knob_value))

        biosxml = get_bios_dump_file_path()
        if biosxml is not None:
            import xmltodict
            with open(biosxml, encoding='utf-8') as f:
                xmlstr = f.read()
                bios_dict = xmltodict.parse(xmlstr)
                assert 'SYSTEM' in bios_dict, f'invalid root in {biosxml}'
                assert 'biosknobs' in bios_dict['SYSTEM'], f'no biosknobs in {biosxml}'
                assert 'knob' in  bios_dict['SYSTEM']['biosknobs'], f'no knobs in biosknobs section'
                for knob in bios_dict['SYSTEM']['biosknobs']['knob']:
                    if knob['@setupType'] == 'ReadOnly':
                        continue
                    name = knob['@name']
                    if 'options' not in knob:
                        value = knob['@CurrentVal']
                        if (name,value) in temp.keys():
                            temp[(name, value)].append((name,value))
                        else:
                            temp[(name, value)] = [(name,value)]
                        continue
                    for opt in knob['options']['option']:
                        if (name, opt['@text']) in temp.keys():
                            temp[(name, opt['@text'])].append((name,opt['@value']))
                        else:
                            temp[(name, opt['@text'])] = [(name, opt['@value'])]
                        if (name, opt['@value']) in temp.keys():
                            temp[(name, opt['@value'])].append((name,opt['@value']))
                        else:
                            temp[(name, opt['@value'])] = [(name, opt['@value'])]

        self.__feature_table = copy.deepcopy(temp)
        #self.__feature_table = sorted(temp.items(), key=lambda x: x[0], reverse=False)

    """
    def get_bios_knob(self, name, value):
        Assert((name, value) in self.feature_table.keys(),
               f'failed to find feature setting [{name} = {value}] in feature table')
        return self.feature_table[(name, value)]
        
    def get_bios_menu_knob(self, name):
        if name in self.bios_menu_knob.keys():
            return self.bios_menu_knob[name]
        else:
            return None
    """

    def load_bios_menu_table(self):
        sheet = self.__get_sheet(self.TAB_NAME_BIOS_MENU)

        temp = {}
        for line in range(2, sheet.max_row + 1):
            name = to_plain_text(sheet.cell(line, 1).value.strip())
            knob_name = to_plain_text(sheet.cell(line, 2).value.strip())

            knob_path = []
            for i in range(3, sheet.max_column + 1):
                v = sheet.cell(line, i).value
                if v is not None and v.strip() != "":
                    item = to_plain_text(v.strip())
                    if item is not None and item != "":
                        knob_path.append(item)

            Assert(name not in temp,
                   f'duplicated definition for bios menu knob: line{line} : name={name}')
            temp[name] = (knob_name, knob_path)

        self.__bios_table = copy.deepcopy(temp)
        #self.__bios_table = sorted(temp.items(), key=lambda x: x[0], reverse=False)


def main():
    mapping_table = MappingTableParser('pvl.xlsx')
    feature_table = mapping_table.feature_table
    print(feature_table)

    feature_pair = {}
    for k, v in feature_table.keys():
        if k not in feature_pair.keys():
            print(k, v)
            feature_pair.update({k: [v]})
        else:
            if k not in feature_pair[k]:
                feature_pair[k].append(v)

    print(feature_pair)





if __name__ == '__main__':
    main()