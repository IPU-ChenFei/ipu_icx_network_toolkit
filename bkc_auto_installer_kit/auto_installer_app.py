#!/usr/bin/env
#################################################################################
# INTEL CONFIDENTIAL
# Copyright Intel Corporation All Rights Reserved.
#
# The source code contained or described herein and all documents related to
# the source code ("Material") are owned by Intel Corporation or its suppliers
# or licensors. Title to the Material remains with Intel Corporation or its
# suppliers and licensors. The Material may contain trade secrets and proprietary
# and confidential information of Intel Corporation and its suppliers and
# licensors, and is protected by worldwide copyright and trade secret laws and
# treaty provisions. No part of the Material may be used, copied, reproduced,
# modified, published, uploaded, posted, transmitted, distributed, or disclosed
# in any way without Intel's prior express written permission.
#
# No license under any patent, copyright, trade secret or other intellectual
# property right is granted to or conferred upon you by disclosure or delivery
# of the Materials, either expressly, by implication, inducement, estoppel or
# otherwise. Any license under such intellectual property rights must be express
# and approved by Intel in writing.
#################################################################################
__author__ = "Naveenraj, Suresh, Monika, Mandar, Kishor"
__version__ = "v3.2"
__maintainer__ = "Naveenraj"
__email__ = "k.naveenraj@intel.com"
__status__ = "Production"

import fileinput
import base64
import os
import ctypes
import subprocess
import tkinter.font as font
import logging
import socket
import time
import paramiko
import zipfile
import pathlib
import sys
import requests
import xml.etree.ElementTree as ET
import threading
import io
import getpass
import urllib.parse
import shutil
import six
import functools
from urllib.parse import urlparse
import six.moves

if six.PY2:
    from pathlib import Path
if six.PY3:
    from pathlib2 import Path

from contextlib import ExitStack
from base64 import b64encode
from datetime import datetime
from subprocess import Popen, PIPE, STDOUT
from tkinter import *
from tkinter import messagebox
from openpyxl import *
from tkinter import ttk


def singleton(cls):
    """
    Singleton decorator

    :param cls: Python class object to enforce as a singleton
    :return: Instance of cls.
    """

    instance = {}

    def get_instance(*args, **kwargs):
        """
        Get the singleton instance of the class, and instantiate it if wasn't already.

        :param args: Positional arguments passed to the constructor of the class
        :param kwargs: Nominal arguments passed to the constructor of the class
        :return: Instance of cls associated with this singleton function
        """
        if cls not in instance.keys():
            instance[cls] = cls(*args, **kwargs)
        return instance[cls]

    return get_instance


def Shared(**kw):
    __instances = dict()
    __keys = kw["key"]
    if isinstance(kw["key"], str):
        __keys = [kw["key"]]

    def _cls(cls):
        @functools.wraps(cls)
        def _wrapper(*args, **kwargs):
            __ser_key = b64encode((":".join(["{}".format(kwargs[k]) for k in __keys])).encode()).decode()
            if __ser_key not in __instances.keys():
                __instances[__ser_key] = cls(*args, **kwargs)
            return __instances[__ser_key]

        return _wrapper

    return _cls


@singleton
class TestContext(ExitStack):
    """Singleton Wrapper of ExitStack."""


def log_cleanup(logger_obj):
    # type: (logging.Logger) -> None
    for handler in logger_obj.handlers:
        handler.close()
        logger_obj.removeHandler(handler)


class CustomFormatter(logging.Formatter):
    """Logging colored formatter"""

    grey = '\x1b[38;21m'
    blue = '\x1b[38;5;39m'
    yellow = '\x1b[38;5;226m'
    red = '\x1b[38;5;196m'
    bold_red = '\x1b[31;1m'
    reset = '\x1b[0m'

    def __init__(self, fmt):
        super().__init__()
        self.fmt = fmt
        self.FORMATS = {
            logging.DEBUG: self.yellow + self.fmt + self.reset,
            logging.INFO: self.blue + self.fmt + self.reset,
            logging.WARNING: self.yellow + self.fmt + self.reset,
            logging.ERROR: self.red + self.fmt + self.reset,
            logging.CRITICAL: self.bold_red + self.fmt + self.reset
        }

    def format(self, record):
        log_fmt = self.FORMATS.get(record.levelno)
        formatter = logging.Formatter(log_fmt)
        return formatter.format(record)


try:
    """
    Fetch the hostname and append to the log for better clarity on where the test executed.
    """
    hostname_res = subprocess.run(["powershell", "-c", "hostname"], check=True, stdout=subprocess.PIPE,
                             stderr=subprocess.PIPE, universal_newlines=True)

    stdout, stderr = hostname_res.stdout, hostname_res.stderr

    host_name = hostname_res.stdout.strip()
    host_name_no_special = ''.join(e for e in host_name if e.isalnum())

    date = datetime.now().strftime("%Y_%m_%d_%I_%M_%S_%p")

    log_dir_name = os.path.join("C:\Auto_Installer_logs")
    if not os.path.exists(log_dir_name):
        os.makedirs(log_dir_name)

    log_file_name = os.path.join(log_dir_name, "{}_auto_installer_{}.log".format(host_name_no_special,date))

    # piv_log_fmt = coloredlogs.install(fmt=u'%(asctime)s,%(msecs)d %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s',
    #                     datefmt=u'%m/%d/%Y %I:%M:%S %p')

    test_log = logging.getLogger()
    test_log.setLevel(logging.DEBUG)
    fmt = u'%(asctime)s,%(msecs)d %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s'
    # Define message format
    piv_log_fmt = logging.Formatter(fmt=fmt, datefmt=u'%m/%d/%Y %I:%M:%S %p')

    # custom_logger_color = CustomFormatter(u'%(asctime)s,%(msecs)d %(levelname)-8s [%(filename)s:%(lineno)d] %(message)s')
    # test_new_log = custom_logger_color.format()

    # Create streams
    stdout_stream = logging.StreamHandler(sys.stdout)
    stdout_stream.setLevel(logging.DEBUG)
    # stdout_stream.setFormatter(piv_log_fmt)
    stdout_stream.setFormatter(CustomFormatter(fmt))
    file_stream = logging.FileHandler(log_file_name, mode='w', encoding='utf-8')
    file_stream.setLevel(logging.DEBUG)
    file_stream.setFormatter(piv_log_fmt)

    # Add handlers to logger
    test_log.addHandler(stdout_stream)
    test_log.addHandler(file_stream)

    TestContext().callback(log_cleanup, test_log)

except Exception as ex:
    raise (ex)


class PduDriver():
    LINUX_PATH_SYS_CONFIG = '/home/Automation/system_configuration.xml'
    WIN_PATH_SYS_CONFIG = 'C:\Automation\system_configuration.xml'

    def __init__(self, test_log):
        if sys.platform != 'win32':
            tree = ET.parse(self.LINUX_PATH_SYS_CONFIG)
        else:
            tree = ET.parse(self.WIN_PATH_SYS_CONFIG)
        root = tree.getroot()

        try:
            for item in root.iter('ac'):
                ip = item.find('./driver/pdu/ip')
                username = item.find('./driver/pdu/username')
                password = item.find('./driver/pdu/password')
                outlet = item.find('./driver/pdu/outlets/outlet')
        except Exception as ex:
            test_log.error("Mismatched Syntax in system_configuration.xml Properly Check the syntax")

        test_log.info("Extracted PDU INFORMATION  ===> {} {} {} {}".format(ip.text, username.text, password.text,
                                                                           outlet.text))
        try:
            ret = subprocess.check_output("ping -n 3 " + ip.text, shell=True)
            if ("0% loss" in str(ret)):
                test_log.info('PDU is alive')
        except Exception as ex:
            test_log.error("PDU is in Turned OFF State or Mentioned Raritan IP is not Correct Ping Failed "
                           "to PDU {0}".format(ip.text))

        try:
            test_log.info("Checking the Credentials of PDU \"{0} {1}\" And Outlet \"{2}\" Status".format(username.text,
                                                                                                           password.text,
                                                                                                           outlet.text))
            self.username = username.text
            self.password = password.text
            self.outlets = outlet.text
            self.port = 22
            self.ip = ip.text
            self.invoke_timeout = 5
            self.powerstate_timeout = 20
            self.cmd_on = ["power outlets {} on /y \n".format(i) for i in self.outlets]
            self.cmd_off = ["power outlets {} off /y \n".format(i) for i in self.outlets]
            self.cmd_show = ["show outlets {} \n".format(self.outlets)]
            self.recv_data = b''
        except Exception as ex:
            test_log.error("PDU Credentials or outlet(s) are not correct, please check and update manually in the "
                           "system configuration xml..")
            test_log.error(ex)

    def get_recv_data(self, ssh):
        self.recv_data = b''
        self.recv_data = ssh.recv(1024)

    def wait_for_invoke(self, ssh):
        nowtime = datetime.now()
        while (datetime.now() - nowtime).seconds < int(self.invoke_timeout):
            t = threading.Thread(target=self.get_recv_data, args=[ssh])
            t.setDaemon(True)
            t.start()
            t.join(3)
            if b'#' in self.recv_data or b'>' in self.recv_data:
                return
        time.sleep(int(self.invoke_timeout))

    def _execute(self, cmd_list):
        ssh = None
        client = None
        try:
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(hostname=self.ip, port=self.port,
                           username=self.username, password=self.password)
            ssh = client.get_transport().open_session()
            ssh.get_pty()
            ssh.invoke_shell()
            self.wait_for_invoke(ssh)
            for cmd in cmd_list:
                num = 0
                while num < 3:
                    ssh.sendall(cmd)
                    time.sleep(0.5)
                    num += 1
        except Exception as ex:
            test_log.error("[{}] {} target failed, the reason is {}".format((datetime.now(), self.ip, str(ex))))
            return False
        finally:
            ssh.close()
            client.close()

    def _check_dict_value(self, dict_data):
        data = set(dict_data.values())
        if len(data) > 1:
            return False
        elif len(data) == 1 and list(data)[0]:
            return True
        elif len(data) == 1 and not list(data)[0]:
            return False
        else:
            return None

    def ac_power_on(self, timeout=None):
        self._execute(self.cmd_on)
        return self.get_ac_power_state(timeout)

    def ac_power_off(self, timeout=None):
        self._execute(self.cmd_off)
        return self.get_ac_power_state(timeout)

    def get_ac_power_state(self, timeout):
        ssh = None
        client = None
        try:
            client = paramiko.SSHClient()
            client.set_missing_host_key_policy(paramiko.AutoAddPolicy())
            client.connect(hostname=self.ip, port=self.port, username=self.username,
                           password=self.password, timeout=timeout)
            ssh = client.get_transport().open_session()
            ssh.get_pty()
            ssh.invoke_shell()
            self.wait_for_invoke(ssh)
            state_list = {}
            for cmd in self.cmd_show:
                num = 0
                while num < 3:
                    ssh.sendall(cmd)
                    time.sleep(0.5)
                    ret_data = ssh.recv(1024).decode("utf-8")
                    if 'On' in ret_data:
                        state_list[cmd] = True
                    elif 'Off' in ret_data:
                        state_list[cmd] = False
                    num += 1
            return self._check_dict_value(state_list)
        except Exception as ex:
            test_log.error("[{}] {} target failed, the reason is {}".format((datetime.now(), self.ip, str(ex))))
        finally:
            ssh.close()
            client.close()


def installer_info():
    """
    Function lists the functionalities that the tools offers.

    :return: None
    """
    from tabulate import tabulate
    table_print = tabulate([["1. Generate System Configuration xml file"], ["2. Cscripts Installation"],
                            ["3. PythonSv Installation (Semi-Automated)"], ["4. Banino software configuration"],
                            ["5. Verify BIOS Serial Communication Port"], ["6. Verify BMC Serial Communication Port"],
                            ["7. Install Sprinter Client"], ["8. Download CC build of DTAF Content Framework"],
                            ["9. Download CC build of DTAF Core Framework"], ["10. Download CC build of Miv-Montana Framework"],
                            ["11. Pycharm IDE Installation"], ["12. Putty Installation"],
                            ["13. WinScp Installation"], ["14. Quartus 18.1 Installation"],
                            ["15. Python 3.6.8 (64 Bit) Installation"], ["16. Verify ITP Probe Connection"],
                            ["17. Install Python Modules"], ["18. Verify Raritan PDU connection"],
                            ["19. Verify BMC (Redfish) connection"], ["20. Install CCB package for Windows"],
                            ["21. Git-Bash Installation"], ["22. Copy Content configuration / Montana Configuration"],
                            ["23. Command Center Instance Change"], ["24. Powercli Installation"],
                            ["25. Background Flask Service"]],
                           headers=['List'], tablefmt="fancy_grid")

    test_log.info("\n+---------------------------------------------+\n"
                  "+           Auto installer Offers             +\n"
                  "+---------------------------------------------+\n{}".format(table_print))


# globally declare wb and sheet variable
sys_config_values = []
is_var_empty = []
sys_config_dict = {}

"""
Remove list variable is to filter the 8S PDU entries shown below, if we add any options before the below three options, 
make sure to increase each numbers in the variable "remove_list" by how many options added.

1. "Second Raritan pdu IP Address or Name (8S only)", 
2. "Second Raritan Pdu First Outlet (8S only)",
3. "Second Raritan Pdu Second Outlet (8S only)"
"""

try:

    config = six.moves.configparser.RawConfigParser()
    config.read(os.path.join(Path(os.path.join(__file__)).parent , "config\\defaults.cfg"))
    env_setup_bat_path = config.get('Paths', 'env_setup_path')
    if os.path.exists(env_setup_bat_path):
        test_log.info("Executing batch file to set the environment variables - {}".format(env_setup_bat_path))
        env_set_res = os.system(env_setup_bat_path)
        if env_set_res == 0:
            test_log.info("Environment variables has been set successfully.")
        else:
            test_log.error("Batch file execution return: {}".format(env_set_res))
    else:
        test_log.error("Batch file not avaliable in {}".format(env_setup_bat_path))
    test_log.info("Starting One Click Auto Installer")
    remove_list = [26, 27, 28]

    sys_config_excel_path = r'C:\Automation\System_configuration.xlsx'
    automation_path = r"C:\Automation"

    atf_username = os.environ['sys_user']
    atf_password = os.environ['sys_pwd']

    dtaf_core_path = None
    dtaf_content_path = None
    platform_chosen = None
    file_name_config_generated = None

    if not os.path.exists(automation_path):
        os.mkdir(automation_path)
        test_log.info("Unable to find 'Automation' folder, so created Automation directory under "
                      "C: drive.. Automation Directory Created")

    try:
        path_dir = r"C:\tools"
        if os.path.exists(path_dir):
            test_log.info(" {} directory exists.. deleting".format(path_dir))
            shutil.rmtree(path_dir)
            test_log.info(" {} directory has been deleted.. creating new one..".format(path_dir))
            os.mkdir(path_dir)
        else:
            test_log.info("{} directory not exists.. creating".format(path_dir))
            os.mkdir(path_dir)
    except NotADirectoryError:
        test_log.info("{} directory not exists.. creating".format(path_dir))
        os.mkdir(path_dir)
    except Exception as ex:
        test_log.error(str(ex))

    wb_create = Workbook()

    wb_create.save(filename=sys_config_excel_path)

    # opening the existing excel file
    wb = load_workbook(sys_config_excel_path)

    # create the sheet object
    sheet = wb.active

    sheet.title = "System_configuration_details"

    label_variables = ["platform_type", "board_name", "platform_name", "cpu_family", "cpu_stepping", "pch_family",
                       "qdf", "sut_os_kernel", "sut_os_name", "sut_os_version", "proxy_link", "cc_build_version",
                       "sut_ip_address", "sut_login_username", "sut_login_password",
                       "bmc_ip_address", "bmc_mac_address", "bios_com_port_number", "bmc_com_port_number",
                       "ladybird_driver_serial_number", "is_this_rasp_system", "raritan_pdu_ip_address_or_name",
                       "raritan_pdu_username", "raritan_pdu_password", "raritan_pdu_first_outlet",
                       "raritan_pdu_second_outlet", "second_raritan_pdu_ip_address_or_name",
                       "second_raritan_pdu_first_outlet", "second_raritan_pdu_second_outlet",
                       "quartus_application_exe_path"]

    labels = ["Platform type", "Board Name", "Platform Name", "CPU Family", "CPU Stepping", "PCH Family", "QDF",
              "SUT OS Kernel Version", "SUT OS Name (Distribution)", "SUT OS Version", "Proxy Link",
              "Command Center DTAF Build Version \n (DPG_Automation_Master_142022_1.0.399_E)",
              "SUT IP Address", "Sut Login Username", "Sut Login Password",
              "Bmc IP Address", "Bmc MAC Address", "BIOS Com Port Number (ex: COM5)", "BMC Com Port Number (ex: COM15)",
              "Ladybird Driver Serial Number", "System Location", "Raritan pdu IP Address or Name",
              "Raritan Pdu Username", "Raritan Pdu Password", "Raritan Pdu First Outlet",
              "Raritan Pdu Second Outlet (if no type NA)", "Second Raritan pdu IP Address or Name (8S only)",
              "Second Raritan Pdu First Outlet (8S only)", "Second Raritan Pdu Second Outlet (8S only)",
              "Quartus Application EXE Path"]
except Exception as ex:
    test_log.error(str(ex))


def excel():
    # resize the width of columns in excel spreadsheet
    try:
        import string

        for letter in string.ascii_uppercase:
            sheet.column_dimensions[letter].width = 30

        # write given data to an excel spreadsheet at particular location
        for column_no in range(1, len(labels) + 1):
            sheet.cell(row=1, column=column_no).value = labels[column_no - 1]
    except Exception as ex:
        test_log.error(str(ex))


# Function for clearing the contents of text entry boxes
def clear():
    """
    Function to clear the content of text entry box

    :return: None
    """
    try:
        for ind in range(1, len(label_variables) + 1):
            var_clear = eval("label_variables[ind-1]")

            if (ind - 1 in remove_list) and platform_chosen == '2S_4S':
                continue
            # globals()[var_clear] --> converts a dynamic eval string to a variable name
            globals()[var_clear].delete(0, END)
    except Exception as ex:
        test_log.error(str(ex))


# Function to take data from GUI window and write to an excel file
def insert(new_window):
    """
    Function to check on un-filled entries.

    :return: None
    """
    try:
        global is_var_empty
        is_var_empty = []

        for ind in range(1, len(label_variables) + 1):
            var_insert = eval("label_variables[ind-1]")

            if (ind - 1 in remove_list) and platform_chosen == '2S_4S':
                continue
            # globals()[var_clear] --> converts a dynamic eval string to a variable name
            if globals()[var_insert].get() == "" or None:
                is_var_empty.append(True)
            else:
                is_var_empty.append(False)

        if any(is_var_empty):

            messagebox.showerror("ERROR !!", "Some or all the inputs are empty, please check and fill all the boxes "
                                             "to proceed further..")

        else:
            # assigning the max row and max column value up to which data is written in an excel sheet to the variable
            current_row = sheet.max_row
            current_column = sheet.max_column

            # get method returns current text as string which we write into excel spreadsheet at particular location
            for ind in range(1, len(label_variables) + 1):
                var_insert = eval("label_variables[ind-1]")
                if (ind - 1 in remove_list) and platform_chosen == '2S_4S':
                    continue
                if not ind - 1 == 11:
                    sheet.cell(row=current_row + 1, column=ind).value = globals()[var_insert].get()
                sys_config_values.append(globals()[var_insert].get())
                sys_config_dict.update({label_variables[ind - 1]: globals()[var_insert].get()})
                test_log.info(sys_config_dict)

            # save the file
            wb.save(sys_config_excel_path)

            # set focus on the name_field box
            globals()[label_variables[0]].focus_set()

            # call the clear() function
            clear()

            messagebox.showinfo("Success", "Form submitted successfully.")

            # Close the window
            new_window.destroy()
    except Exception as ex:
        test_log.error(str(ex))


def generate_system_config_xml(platform_chosen, cpld_application_path=None):
    try:
        import xml.etree.ElementTree as ET
        config_file_name = None

        sys_path = Path(os.path.join(__file__)).parent

        if "2S_4S" in platform_chosen:
            config_file_name = os.path.join(sys_path, "system_configuration_2S_4S.xml")
        elif "8S" in platform_chosen:
            config_file_name = os.path.join(sys_path, "system_configuration_8S.xml")
        else:
            test_log.error("Given platform is not supported, please provide between 2S, 4S or 8S.. Platform Error !!! ")
            exit()

        tree = ET.parse(config_file_name)
        root = tree.getroot()

        # using bool() check if dictionary is empty
        # res_dict = not bool(sys_config_dict)
        temp_dict = 0
        for key, values in sys_config_dict.items():
            if sys_config_dict[key] == "" or None:
                temp_dict = temp_dict + 1

        if temp_dict == len(sys_config_dict):
            ctypes.windll.user32.MessageBoxW(0, "Exiting the form..", "Exit !! ", 0)
            exit()

        # modify platform type
        for elem in root.iter('platform'):
            elem.set('type', sys_config_dict["platform_type"])

        # modify platform Name
        for elem in root.iter('platform'):
            elem.set('name', sys_config_dict["platform_name"])

        # modify Board name
        for elem in root.iter('platform'):
            elem.set('boardname', sys_config_dict["board_name"])

        # modify cpu_family
        for elem in root.findall("suts/sut/silicon/cpu/family"):
            elem.text = sys_config_dict["cpu_family"]

        # modify cpu_stepping
        for elem in root.findall("suts/sut/silicon/cpu/stepping"):
            elem.text = sys_config_dict["cpu_stepping"]

        # modify pch_family
        for elem in root.findall("suts/sut/silicon/pch/family"):
            elem.text = sys_config_dict["pch_family"]

        # modify qdf
        for elem in root.findall("suts/sut/silicon/cpu/qdf"):
            elem.text = sys_config_dict["qdf"]

        # modifying an attribute -- SUT IP Address
        for elem in root.iter('sut'):
            elem.set('ip', sys_config_dict["sut_ip_address"])

        # changing a field text -- Adding BMC ip
        for elem in root.iter('ip'):
            elem.text = sys_config_dict["bmc_ip_address"]

        # modifying an attribute -- Adding SUT type sut_os name attribute
        for elem in root.iter('sut_os'):
            elem.set('version', sys_config_dict["sut_os_version"])
            elem.set('kernel', sys_config_dict["sut_os_kernel"])
            if str(sys_config_dict["sut_os_name"]).lower() == "rhel" or \
                    str(sys_config_dict["sut_os_name"]).lower() == "centos" or \
                    str(sys_config_dict["sut_os_name"]).lower() == "esxi" or \
                    str(sys_config_dict["sut_os_name"]).lower() == "cent os":
                elem.set('name', 'Linux')
            else:
                elem.set('name', 'Windows')

        # modifying an attribute -- Adding SUT OS subtype
        for elem in root.iter('sut_os'):
            elem.set('subtype', sys_config_dict["sut_os_name"])

        # modifying an attribute -- Adding SUT username and password
        for elem_cre in root.findall("suts/sut/providers/sut_os/driver/ssh/credentials"):
            elem_cre.set('user', sys_config_dict["sut_login_username"])
            elem_cre.set('password', sys_config_dict["sut_login_password"])

        # changing a field text -- Adding SUT IP
        for elem in root.iter('ipv4'):
            elem.text = sys_config_dict["sut_ip_address"]

        # changing a field text -- Adding COM port
        for elem_port in root.findall('suts/sut/providers/console_log/driver/com/port'):
            elem_port.text = sys_config_dict["bios_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/uefi_shell/driver/com/port'):
            elem_port.text = sys_config_dict["bios_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/bios_setupmenu/driver/com/port'):
            elem_port.text = sys_config_dict["bios_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/bios_bootmenu/driver/com/port'):
            elem_port.text = sys_config_dict["bios_com_port_number"]

        # changing a field text -- Adding COM port
        for elem_port in root.findall('suts/sut/providers/console/driver/com/port'):
            elem_port.text = sys_config_dict["bmc_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/silicon/bmc/bmcport'):
            elem_port.text = sys_config_dict["bmc_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/flash/driver/redfish/bmcport'):
            elem_port.text = sys_config_dict["bmc_com_port_number"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/physical_control/driver/redfish/ip'):
            elem_port.text = sys_config_dict["bmc_ip_address"]

        # changing a field text
        for elem_port in root.findall('suts/sut/providers/physical_control/driver/redfish/mac'):
            elem_port.text = sys_config_dict["bmc_mac_address"]

        # changing a field text -- Adding the XMLCI folder path according to the OS type
        for elem_sutospath in root.iter('sutospath'):
            if str(sys_config_dict["sut_os_name"]).lower() == "Windows".lower() or \
                str(sys_config_dict["sut_os_name"]).lower() == "Win".lower():
                elem_sutospath.text = r'C:\BKCPkg\xmlcli\\'
            else:
                elem_sutospath.text = 'opt/APP/xmlcli/'

        # changing a field text -- Adding the SUT IP for xmlcli
        for elem_ip in root.findall('suts/sut/providers/bios/driver/xmlcli/ip'):
            elem_ip.text = sys_config_dict["sut_ip_address"]

        # changing a field text -- Adding the SUT user name for xmlci
        for elem_user in root.findall('suts/sut/providers/bios/driver/xmlcli/user'):
            elem_user.text = sys_config_dict["sut_login_username"]

        # changing a field text -- Adding the SUT password for xmlci
        for elem_pass in root.findall('suts/sut/providers/bios/driver/xmlcli/password'):
            elem_pass.text = sys_config_dict["sut_login_password"]

        for elem in root.iter('ladybird_driver_serial'):
            elem.text = sys_config_dict["ladybird_driver_serial_number"]

        for elem in root.iter('rasp'):
            if str(sys_config_dict["is_this_rasp_system"]).lower() == "flex lab" or \
                str(sys_config_dict["is_this_rasp_system"]).lower() == "rasp lab":
                elem.text = "True"
            else:
                elem.text = "False"

        if platform_chosen == "2S_4S":
            tag_path_outlet = 'suts/sut/providers/ac/driver/pdu/outlets'
            tag_path_ip_add = 'suts/sut/providers/ac/driver/pdu/ip'
            tag_path_ip_username = 'suts/sut/providers/ac/driver/pdu/username'
            tag_path_ip_pass = 'suts/sut/providers/ac/driver/pdu/password'
        else:
            tag_path_outlet = 'suts/sut/providers/ac/driver/pdu/groups'
            tag_path_ip_add = 'suts/sut/providers/ac/driver/pdu/groups'
            tag_path_ip_username = 'suts/sut/providers/ac/driver/pdu/groups/group/username'
            tag_path_ip_pass = 'suts/sut/providers/ac/driver/pdu/groups/group/password'

        if platform_chosen == "8S":
            for elem_pdu_ip in root.findall(tag_path_ip_add):
                for elem_pdu_ip_inner in elem_pdu_ip.findall("group")[0]:
                    if elem_pdu_ip_inner.tag == "ip":
                        elem_pdu_ip_inner.text = sys_config_dict["raritan_pdu_ip_address_or_name"]
                        break
                for elem_pdu_ip_inner in elem_pdu_ip.findall("group")[1]:
                    if elem_pdu_ip_inner.tag == "ip":
                        elem_pdu_ip_inner.text = sys_config_dict["second_raritan_pdu_ip_address_or_name"]
                        break
        else:
            for elem_pdu_ip in root.findall(tag_path_ip_add):
                elem_pdu_ip.text = sys_config_dict["raritan_pdu_ip_address_or_name"]

        for elem_pdu_user in root.findall(tag_path_ip_username):
            elem_pdu_user.text = sys_config_dict["raritan_pdu_username"]

        for elem_pdu_pass in root.findall(tag_path_ip_pass):
            elem_pdu_pass.text = sys_config_dict["raritan_pdu_password"]

        if platform_chosen == "8S":
            for elem_pdu_group in root.findall(tag_path_outlet):
                for elem_pdu_outlet in elem_pdu_group.findall("group")[0]:
                    if elem_pdu_outlet.tag == "outlets":
                        elem_pdu_outlet[0].text = sys_config_dict["raritan_pdu_first_outlet"]
                        elem_pdu_outlet[1].text = sys_config_dict["raritan_pdu_second_outlet"]

                for elem_pdu_8s_outlet in elem_pdu_group.findall("group")[1]:
                    if elem_pdu_8s_outlet.tag == "outlets":
                        elem_pdu_8s_outlet[0].text = sys_config_dict["second_raritan_pdu_first_outlet"]
                        elem_pdu_8s_outlet[1].text = sys_config_dict["second_raritan_pdu_second_outlet"]
        else:
            for elem_out_1 in root.findall(tag_path_outlet):
                elem_out_1[0].text = sys_config_dict["raritan_pdu_first_outlet"]
                # no of pdu outlets
                if str(sys_config_dict["raritan_pdu_second_outlet"]) == "NA":
                    elem_out_1.remove(elem_out_1[1])
                else:
                    elem_out_1[1].text = sys_config_dict["raritan_pdu_second_outlet"]

        test_log.info("cpld_application_path---> {}".format(cpld_application_path))
        for elem in root.iter('cpld_application_path'):
            elem.text = cpld_application_path

        if cpld_application_path is None:
            for elem in root.iter('cpld_application_path'):
                elem.text = sys_config_dict["quartus_application_exe_path"]

        global file_name_config_generated

        file_name_config_generated = "C:\\Automation\\system_configuration.xml"

        date = datetime.now().strftime("%Y_%m_%d_%I_%M_%S_%p")
        renamed_file = file_name_config_generated.split(".")[0] + "_previous_" + date

        if os.path.isfile(file_name_config_generated):
            cmd_rename = "Rename-Item -Path {} -NewName {}.xml".format(file_name_config_generated, renamed_file)

            test_log.info(cmd_rename)
            cmd_res = subprocess.run(["powershell", "-c", cmd_rename], check=True, stdout=subprocess.PIPE,
                                     stderr=subprocess.PIPE, universal_newlines=True)

            stdout, stderr = cmd_res.stdout, cmd_res.stderr

            if cmd_res.returncode != 0:
                test_log.error("An error occurred: {}".format(stderr))
            else:
                test_log.info("Command executed successfully, previous system configuration file renamed "
                              "to '{}'..".format(renamed_file))

        test_log.info("Writing new system configuration file..")
        tree.write(file_name_config_generated, encoding='utf-8', xml_declaration=True)
        test_log.info("Writing system configuration file is successful..")

        test_log.info("System configuration xml file generated successfully..")

    except Exception as ex:
        test_log.error("Not able to generate the system configuration xml file due to {}".format(ex))


def get_platform_type():
    return platform_chosen


def open_platform_based_window(button_pressed):
    try:
        temp_ind = 0
        temp_ind_box = 0
        increase_row_pos_by = 1

        # Close the window
        # root.destroy()

        global platform_chosen
        platform_chosen = button_pressed

        # Toplevel object which will be treated as a new window
        new_window = Tk()

        new_window.state("zoomed")

        # set the background colour of GUI window
        new_window.configure(background='light blue')

        # set the title of GUI window
        new_window.title("Generate System Configuration file Form")

        # set the configuration of GUI window
        new_window.geometry("750x650")

        excel()

        # create a Form label
        heading = Label(new_window, text="System configuration Form", bg='light blue', font=('Helvetica', 25))
        heading.grid(row=0, column=0, columnspan=10, pady=15)

        heading_sub = Label(new_window, text="*** Please fill in all the boxes below before submitting ***",
                            bg='Green', fg='white')
        heading_sub.grid(row=1, column=0, columnspan=8, pady=10)

        for index in range(1, len(label_variables) + 1):
            variable = eval("label_variables[index-1]")

            # locals()[variable] --> converts a dynamic eval string to a variable name
            if (index - 1 in remove_list) and platform_chosen == '2S_4S':
                pass
            else:
                # create labels for each entry box
                globals()[variable] = Label(new_window, text=labels[index - 1], bg='light blue')

        for index in range(1, len(label_variables) + 1):
            variable = eval("label_variables[index-1]")

            if (index - 1 in remove_list) and platform_chosen == '2S_4S':
                pass
            else:
                # grid method is used for placing the widgets at respective positions in table like structure.
                # locals()[variable] --> converts a dynamic eval string to a variable name
                if index < round(len(label_variables) / 2):
                    globals()[variable].grid(row=index + increase_row_pos_by, column=0, padx=5, pady=5)
                else:
                    temp_ind = temp_ind + 1
                    globals()[variable].grid(row=temp_ind + increase_row_pos_by, column=3, padx=5, pady=5)

            # create a text entry box for typing the information
            # if index - 1 == 11:
            #     globals()[variable] = Entry(new_window, show="*", justify="center")
            #     # grid method is used for placing the widgets at respective positions in table like structure .
            #     globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="55", padx=5, pady=5)

            if index - 1 == 0:
                # Options you want to add
                choices_type = ['reference', 'dell', 'hp', 'microsoft', 'lenovo']

                globals()[variable] = ttk.Combobox(new_window, value=choices_type, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)
            elif index - 1 == 1:
                # Options you want to add
                choices_boardname = ['ArcherCity', 'WilsonCity', 'NeonCity']

                globals()[variable] = ttk.Combobox(new_window, value=choices_boardname, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)
            elif index - 1 == 2:
                # Options you want to add
                choices_platform_name = ['SappahireRapids', 'EmeraldRapids']

                globals()[variable] = ttk.Combobox(new_window, value=choices_platform_name, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)
            elif index - 1 == 3:
                # Options you want to add
                choices_cpu_family = ['SPR', 'EMR']

                globals()[variable] = ttk.Combobox(new_window, value=choices_cpu_family, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)
            elif index - 1 == 4:
                # Options you want to add
                choices_cpu_stepping = ['A0', 'A1', 'A2', 'B0', 'B1', 'B2', 'C0', 'C1', 'C2', 'D0', 'E0', 'E1', 'E2',
                                        'E3', 'E4', 'E5', 'E6']

                globals()[variable] = ttk.Combobox(new_window, value=choices_cpu_stepping, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)
            elif index - 1 == 5:
                # Options you want to add
                choices_pch_family = ['EBG', 'EMT']

                globals()[variable] = ttk.Combobox(new_window, value=choices_pch_family, justify="center")
                globals()[variable].current(0)
                # grid method is used for placing the widgets at respective positions in table like structure .
                globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="45", padx=5, pady=5)

            elif (index - 1 in remove_list) and platform_chosen == '2S_4S':
                pass
            else:
                globals()[variable] = Entry(new_window, justify="center")

                # grid method is used for placing the widgets at respective positions in table like structure .
                if index < round(len(label_variables) / 2):
                    globals()[variable].grid(row=index + increase_row_pos_by, column=1, ipadx="55", padx=5, pady=5)
                else:
                    temp_ind_box = temp_ind_box + 1
                    if index - 1 == 20:
                        # Options you want to add
                        choices_system = ['Local lab (IIND)', 'RASP Lab', 'FLEX Lab']

                        globals()[variable] = ttk.Combobox(new_window, value=choices_system, justify="center")
                        globals()[variable].current(0)
                        # grid method is used for placing the widgets at respective positions in table like structure .
                        globals()[variable].grid(row=temp_ind_box + increase_row_pos_by, column=4, ipadx="45", padx=5,
                                                 pady=5)
                    else:
                        globals()[variable].grid(row=temp_ind_box + increase_row_pos_by, column=4, ipadx="55", padx=5,
                                             pady=5)

        globals()[label_variables[10]].insert(0, "http://proxy01.iind.intel.com:912")

        # call excel function
        excel()

        # define font
        myfont = font.Font(family='Helvetica')

        # create a Submit Button and place into the new_window window
        submit = Button(new_window, text="Submit", font=myfont, command=lambda: insert(new_window), bg='#0052cc',
                        fg='#ffffff')
        submit.grid(row=len(label_variables) + 2, column=3, padx=15, pady=5, ipadx=10)
    except Exception as ex:
        ctypes.windll.user32.MessageBoxW(0, str(ex), "Exception !!!", 0)


def cc_build_download():
    try:
        cwd = r"C:\tools"
        curl = "https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/cc-client.zip"
        cc_client_zip_name = "\\" + curl.split("/")[-1]
        cc_client_name = "\\" + cc_client_zip_name.split(".zip")[0]

        time.sleep(5)

        opt = subprocess.Popen(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET " + str(curl) + " --output " + str(cwd) + cc_client_zip_name, stdin=PIPE,
            stdout=PIPE, stderr=STDOUT, shell=True)

        opt = opt.stdout.read()
        if str(opt).find(r"100") != -1:
            test_log.info(opt)
            test_log.info("{} file downloaded from artifactory successfully..".format(cc_client_zip_name))
        else:
            test_log.error("{} download from artifactory Failed..".format(cc_client_zip_name))
            return False

        path_to_zip_file = cwd + cc_client_zip_name
        path_to_extract_zip = cwd + cc_client_name
        with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
            zip_ref.extractall(path_to_extract_zip)
            zip_ref.close()
        test_log.info("CC-Client Tool Extraction Successful")

        dtaf_path = "C:\\DPG_Automation"

        if os.path.exists(dtaf_path):
            ret = os.system("RMDIR /S /Q {}".format(dtaf_path))
            if ret == 0:
                test_log.info("Successfully removed the existing '{}' directory and its contents..".format(dtaf_path))
        else:
            test_log.info("'{}' directory does not exists... Creating one..".format(dtaf_path))

        os.mkdir(dtaf_path)

        test_log.info("'{}' directory created..".format(dtaf_path))

        # example --> cc_build_version = DPG_Automation_testing_bkc_229_fixes_142022_1.0.399_E
        cc_test_build_version = sys_config_dict["cc_build_version"]

        cc_user = os.environ["sys_bkc_exec_user"]
        cc_pwd = os.environ["sys_bkc_exec_pass"]

        if cc_test_build_version:
            test_log.info("Waiting for the Command Center to download the frameworks onto the mapped Host..")

            ret = subprocess.check_output(
                r"C:\tools\cc-client\CCClient.exe download framework --project BKC-EGS --instance IIND --frameworkName DTAF --version {} --username ".format(cc_test_build_version) + str(cc_user) + " --password " + str(cc_pwd) + " --path {}".format(dtaf_path), shell=True)

            if "successfully downloaded" in str(ret):
                test_log.info("DTAF / Montana frameworks downloaded successfully..")
            else:
                test_log.error("DTAF / Montana frameworks download Failed..")
                return False

            build_to_zip_file = os.path.join(dtaf_path, cc_test_build_version + ".zip")
            path_to_extract_zip = cwd + cc_client_name
            with zipfile.ZipFile(build_to_zip_file, 'r') as zip_ref:
                zip_ref.extractall(dtaf_path)
                zip_ref.close()
            test_log.info("{} Build Extraction Successful".format(cc_test_build_version))
            return True
        else:
            test_log.info("Command center build version is not given, please check and provide it appropriately..")
            return False
    except Exception as ex:
        test_log.error(ex)
        return False


def copy_config_from_dtaf_content_framework(config_name_in_framework):
    try:
        config_file_src_path = None
        domain_config_name = config_name_in_framework
        config_file_path = r'C:\DPG_Automation\dtaf_content\src\configuration'
        for root, dirs, files in os.walk(str(config_file_path)):
            for name in files:
                if name.startswith(domain_config_name.split(".")[0]) and name.endswith(domain_config_name):
                    config_file_src_path = os.path.join(root, name)

        if config_file_src_path is None:
            test_log.info("{} file does not exists under C:/Automation directory or inside of "
                          "framework, please check..".format(config_name_in_framework))
            return False

        cwd = r"C:\Automation"
        opt = subprocess.Popen(
            r"xcopy /y /f {} "r"{}".format(config_file_src_path, cwd), cwd=cwd, stdin=PIPE, stdout=PIPE,
            stderr=STDOUT, shell=True)

        opt = opt.stdout.read()

        if str(opt).find(r"1 File(s) copied") != -1:
            test_log.info("{} copied to '{}' successfully..".format(config_name_in_framework, cwd))

        if "miv" in config_name_in_framework:
            cmd_res = subprocess.run(["powershell", "-c", "hostname"], check=True, stdout=subprocess.PIPE,
                                     stderr=subprocess.PIPE, universal_newlines=True)

            stdout, stderr = cmd_res.stdout, cmd_res.stderr

            if cmd_res.returncode != 0:
                test_log.error("An error occurred: {}".format(stderr))
            else:
                test_log.info("Powershell Command executed successfully, host name found.. \n{}".format(stdout))
            file_exists_path = cwd + "\\" + stdout.strip() + ".cfg"

            test_log.info("File path : {}".format(file_exists_path))

            if os.path.isfile(file_exists_path):
                test_log.info("miv cfg file already exists with renamed to hostname under '{}'.."
                              "".format(file_exists_path))
            else:
                cmd_rename = "Rename-Item -Path {} -NewName {}.cfg".format(cwd + "\\" + config_name_in_framework,
                                                                           cwd + "\\" + stdout.strip())

                cmd_res = subprocess.run(["powershell", "-c", cmd_rename], check=True, stdout=subprocess.PIPE,
                                         stderr=subprocess.PIPE, universal_newlines=True)

                stdout, stderr = cmd_res.stdout, cmd_res.stderr

                if cmd_res.returncode != 0:
                    test_log.error("An error occurred: {}".format(stderr))
                else:
                    test_log.info("Powershell Command executed successfully, miv cfg file renamed to '{}'.."
                                  "".format(file_exists_path))
    except Exception as ex:
        test_log.error(ex)


def verify_com_port(comport_list):
    """
    Function to verify the COM ports entered by the user is available in the Device Manager

    :param comport_list: list of ports can be given
    :return: True / False
    """

    try:
        import subprocess
        import re

        powershell_cmd = "Get-WmiObject -query 'SELECT Name FROM Win32_PnPEntity' | " \
                         "Where {$_.Name -Match 'COM\d+'} | findstr COM"
        cmd_res = subprocess.run(["powershell", "-c", powershell_cmd], check=True, stdout=subprocess.PIPE,
                                 stderr=subprocess.PIPE, universal_newlines=True)

        stdout, stderr = cmd_res.stdout, cmd_res.stderr

        if cmd_res.returncode != 0:
            test_log.error("An error occurred: {}".format(stderr))
        else:
            test_log.info("Powershell Command executed successfully, Com ports are found.. \n{}".format(stdout))

        comport_list_upp = [x.upper() for x in list(comport_list.split(" "))]

        if comport_list is None:
            test_log.error("Com ports are not given, please check..")

        comport_device_manager_list = re.findall(r"(COM.\d*)", stdout)

        test_log.info("List of com ports given by the user: {}".format(comport_list_upp))
        test_log.info("List of com ports from device manager: {}".format(comport_device_manager_list))

        if all(y in comport_device_manager_list for y in comport_list_upp):
            test_log.info("COM PORTS are matched successfully ..")
        else:
            test_log.error("Com ports are not matching, please check..")

    except Exception as ex:
        test_log.error(ex)


def check_bios_serial():
    import serial
    serial_output_list = []

    ser = serial.Serial('COM10', 115200)
    test_log.info("1")
    ser.write(bytes("\n", encoding='utf8'))
    test_log.info("2")
    serial_output_list.append(ser.readline(1000))
    test_log.info("3")
    ser.write(b'root')
    ser.flush()
    time.sleep(3)
    serial_output_list.append(ser.readline(1000))
    test_log.info("4")
    ser.write(bytes("\n", encoding='utf8'))
    ser.flush()
    time.sleep(3)
    serial_output_list.append(ser.readline(1000))
    test_log.info("5")
    ser.write(b'0penBmc1')
    ser.flush()
    time.sleep(3)
    serial_output_list.append(ser.readline(1000))
    test_log.info("6")
    ser.write(bytes("\n", encoding='utf8'))
    time.sleep(3)
    serial_output_list.append(ser.readline(1000))
    test_log.info("7")
    serial_output_list.append(ser.readline(1000))
    ser.close()

    test_log.info(serial_output_list)

    for op in serial_output_list:
        if "Login" or "Password:" or "Incorrect Login" in op:
            test_log.info("This COM port is for BMC console")
            break


def set_nuc_proxy_repo_clone():
    try:
        proxy = None
        proxy = sys_config_dict["proxy_link"]

        if proxy is None:
            proxy = "http://proxy01.iind.intel.com:912"
            test_log.info("Proxy not provided in the GUI Form.. \nProceeding with default one : {}".format(proxy))
        else:
            test_log.info("Proceeding with proxy : {}".format(proxy))

        git_path = "C:\\Program Files\\Git\\bin\\".replace(" ","` ")

        # setting up nuc proxy
        proxy_github_cmd = "{}\git config --global http.https://github.com.proxy {}".format(git_path, proxy)
        proxy_lfs_github_cmd = "{}\git config --global http.https://lfs.github.com.proxy {}".format(git_path, proxy)
        proxy_cloud_aws_github_cmd = "{}\git config --global http.https://github-cloud.s3.amazonaws.com.proxy {}".format(
            git_path, proxy)
        proxy_cloud_user_content_github_cmd = "{}\git config --global http.https://github-cloud.githubusercontent." \
                                              "com.proxy {}".format(git_path, proxy)

        process = subprocess.run(["powershell", "-c",proxy_github_cmd], check=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        # test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("Github proxy set successfully on NUC.")
        else:
            test_log.info(process.stderr)
            return False

        process = subprocess.run(["powershell", "-c", proxy_lfs_github_cmd], check=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        # test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("LFS Github proxy set successfully on NUC.")
        else:
            test_log.info(process.stderr)
            return False

        process = subprocess.run(["powershell", "-c", proxy_cloud_aws_github_cmd], check=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        # test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("Cloud AWS Github proxy set successfully on NUC.")
        else:
            test_log.info(process.stderr)
            return False

        process = subprocess.run(["powershell", "-c", proxy_cloud_user_content_github_cmd], check=True,
                                 stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)
        # test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("Cloud User Content Github proxy set successfully on NUC.")
        else:
            test_log.info(process.stderr)
            return False
    except Exception as ex:
        test_log.error(ex)


def install_python_modules():
    try:
        python_version = "python-3.8.10-amd64.exe"
        python_path = "C:\\Python38"
        python_path_scripts = "C:\\Python38\\Scripts"

        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_version = "python-3.8.10-amd64.exe"
        #     python_path = "C:\\Python38"
        #     python_path_scripts = "C:\\Python38\\Scripts"
        # else:
        #     python_version = "python-3.6.8-amd64.exe"
        #     python_path = "C:\\Python36"
        #     python_path_scripts = "C:\\Python36\\Scripts"

        dtaf_content_path = r"C:\DPG_Automation\dtaf_content"
        dtaf_core_path = r"C:\DPG_Automation\dtaf_core"
        proxy = sys_config_dict["proxy_link"]
        cmd = "{}\pip install --proxy={} -r {}"

        #ensure pip
        ensure_pip_cmd = "{}\python -m ensurepip".format(python_path)
        process = subprocess.run(ensure_pip_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)
        test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("pip is available")
        else:
            test_log.error(process.stderr)

        # updating pip
        update_pip_cmd = "{}\python -m pip install --proxy={} --upgrade pip".format(python_path, proxy)
        process = subprocess.run(update_pip_cmd, creationflags=subprocess.CREATE_NEW_CONSOLE, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)
        test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("pip is updated")
        else:
            test_log.error(process.stderr)

        # install dtaf_content requirements
        requirement_file = dtaf_content_path + r"\requirements_py3.txt"
        dtaf_content_cmd = cmd.format(python_path_scripts, proxy, requirement_file)
        process = subprocess.run(dtaf_content_cmd, creationflags=subprocess.CREATE_NEW_CONSOLE, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)
        test_log.info(process.stdout)
        # test_log.info(process.stderr)
        if process.returncode == 0:
            test_log.info("Required modules for dtaf_content are installed successfully")
        else:
            test_log.error(process.stderr)

        # install dtaf_core requirements
        requirement_file = dtaf_core_path + r"\requirements\requirements_st_py37.txt"
        dtaf_core_cmd = cmd.format(python_path_scripts, proxy, requirement_file)

        requirement_file_ut = dtaf_core_path + r"\requirements\requirements_ut_py37.txt"
        dtaf_core_cmd_ut = cmd.format(python_path_scripts, proxy, requirement_file_ut)
        commands = [dtaf_core_cmd, dtaf_core_cmd_ut]
        for command in commands:
            process = subprocess.run(command, creationflags=subprocess.CREATE_NEW_CONSOLE, stdout=PIPE, stderr=PIPE, shell=True)
            test_log.info(process.returncode)
            test_log.info(process.stdout)
            # test_log.info(process.stderr)
            if process.returncode == 0:
                test_log.info("Required modules for dtaf_core are installed successfully")
            else:
                test_log.error(process.stderr)
    except Exception as ex:
        test_log.error(ex)


def update_reg_sz_env_path(program_path):

    try:
        import winreg
        import ctypes

        with winreg.ConnectRegistry(None, winreg.HKEY_LOCAL_MACHINE) as hkey:  # Get the necessary HKEY
            with winreg.OpenKey(hkey, "SYSTEM\CurrentControlSet\Control\Session Manager\Environment", 0,
                                winreg.KEY_ALL_ACCESS) as sub_key:  # Go to the environment key
                existing_path_value = winreg.QueryValueEx(sub_key, "PATH")
                test_log.info(existing_path_value)
                new_path_value = existing_path_value[
                                     0] + ";" + program_path + ";"  # Takes the current path value and appends the new program path
                test_log.info(new_path_value)
                winreg.SetValueEx(sub_key, "PATH", 0, winreg.REG_EXPAND_SZ,
                                  new_path_value)  # Updated the path with the updated path

                # Tell other processes to update their environment
                HWND_BROADCAST = 0xFFFF
                WM_SETTINGCHANGE = 0x1A
                SMTO_ABORTIFHUNG = 0x0002
                result = ctypes.c_long()
                SendMessageTimeoutW = ctypes.windll.user32.SendMessageTimeoutW
                SendMessageTimeoutW(HWND_BROADCAST, WM_SETTINGCHANGE, 0, u"Environment", SMTO_ABORTIFHUNG, 5000,
                                    ctypes.byref(result),)

    except Exception as ex:
        test_log.error(ex)
        return False


def add_reg_sz_key_env_path(pathname, path_values_list):
    try:
        import winreg

        if isinstance(path_values_list, list):
            path_values = (";".join(path_values_list)) + ";"
        else:
            path_values = path_values_list
        # test_log.info("System Env pathname : {}".format(pathname))
        # test_log.info("System Env path values : {}".format(path_values))

        keyval = r"SYSTEM\CurrentControlSet\Control\Session Manager\Environment\PYTHONPATH"
        if not os.path.exists(keyval):
            key = winreg.CreateKey(winreg.HKEY_LOCAL_MACHINE, keyval)
        Registrykey = winreg.OpenKey(winreg.HKEY_LOCAL_MACHINE,
                                     r"SYSTEM\CurrentControlSet\Control\Session Manager\Environment", 0,
                                     winreg.KEY_ALL_ACCESS)
        winreg.SetValueEx(Registrykey, pathname, 0, winreg.REG_SZ, path_values)
        test_log.info("{} has been set successfully.".format(pathname))
        test_log.info("Path that was set is : {}".format(winreg.QueryValueEx(Registrykey, pathname)))
        winreg.CloseKey(Registrykey)

    except WindowsError as ex:
        test_log.error(ex)


def dot_net_framework_3_5_installation():
    try:
        cwd = r"C:\tools"
        install_dotnet_cmd_online = "Dism /online /Enable-Feature /FeatureName:'Netfx3'"
        # installing .net 3.5 via online
        process_online = subprocess.run(install_dotnet_cmd_online, stdout=PIPE, stderr=PIPE, shell=True)
        # test_log.info(process_online.returncode)
        if process_online.returncode != 0:
            test_log.info(".Net Framework 3.5 did not install from online")
            test_log.info("Proceeding offline method..")
            try:
                test_log.info(".Net Framework 3.5 is downloading from artifactory..")
                subprocess.check_output(
                    "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/dotnetfx35.exe --output C:\\tools\dotnetfx35.exe",
                    shell=True)
                test_log.info(".Net Framework 3.5 successfully downloaded from artifactory..")

            except Exception as ex:
                test_log.error(".Net Framework 3.5 failed to download..")
                return False

            # installing .net 3.5 via offline
            install_dotnet_cmd_offline = r"C:\tools\dotnetfx35.exe /q /norestart"

            process = subprocess.run(install_dotnet_cmd_offline, stdout=PIPE, stderr=PIPE, shell=True)
            test_log.info(process.returncode)
            if process.returncode == 0:
                test_log.info(".Net Framework 3.5 installed successfully")
            return True
    except Exception as ex:
        test_log.error(ex)


def install_sprinter_atf():
    import zipfile

    try:
        cwd = r"C:\tools"
        curl = "https://af01p-igk.devtools.intel.com/artifactory/testercontroller-igk-local/Sprinter/release/" \
               "Sprinter-v1.1.27.zip"
        sprinter_zip_name = "\\" + curl.split("/")[-1]
        sprinter_name = "\\" + sprinter_zip_name.split(".zip")[0]

        time.sleep(5)
        opt = subprocess.Popen(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET " + str(curl) + " --output " + str(cwd) + sprinter_zip_name, stdin=PIPE,
            stdout=PIPE, stderr=STDOUT, shell=True)

        opt = opt.stdout.read()
        if str(opt).find(r"100") != -1:
            test_log.info(opt)
            test_log.info("{} file downloaded from artifactory successfully..".format(sprinter_zip_name))
        else:
            test_log.error("{} download from artifactory Failed..".format(sprinter_zip_name))
            return False

        time.sleep(3)
        dot_net_res = dot_net_framework_3_5_installation()
        if not dot_net_res:
            test_log.info("retrying one more time..")
            dot_net_framework_3_5_installation()

        path_to_zip_file = cwd + sprinter_zip_name
        path_to_extract_zip = cwd + sprinter_name
        with zipfile.ZipFile(path_to_zip_file, 'r') as zip_ref:
            zip_ref.extractall(path_to_extract_zip)
            zip_ref.close()
        test_log.info("Sprinter Tool Extraction Successful")

        app_config_file_path = os.path.join(path_to_extract_zip, "windows-x64\Sprinter\AppConfig.yml")

        # set sprinter instance to Bangalore..
        sprinter_instance_change(app_config_file_path)

        auto_sprinter_startup_cmd = path_to_extract_zip + "\Setup.exe --path C:\Sprinter\CommandCenter " \
                                                          "--runAsService --closeInstallerWhenFinished"

        # installing Sprinter as windows service
        process = subprocess.run(auto_sprinter_startup_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)
        if process.returncode == 0:
            test_log.info("Sprinter installed successfully and added as windows service.")



        return True
    except Exception as ex:
        test_log.error(ex)


def sprinter_instance_change(app_config_file_path):
    """
    Function to change the instance of the printer to Bangalore location.

    :return: True if instance changed to Bangalore else False
    """

    import subprocess
    import psutil
    import os
    import yaml

    SPRINTER_PATH = "c:\Sprinter\CommandCenter\Sprinter"
    PROCNAME = "Sprinter.exe"
    CC_INSTANCE_NAME = 'Bangalore'

    try:
        # Kill existing sprinter session
        for proc in psutil.process_iter():
            # check whether the process name matches
            if proc.name() == PROCNAME:
                proc.kill()
    except Exception as ex:
        test_log.error("Unable to kill the sprinter process running in the background due to : {}".format(ex))
        return False

    # Update Commandcnetre instance
    with io.open(app_config_file_path, "r+") as file:
        try:
            data = yaml.safe_load(file)
            for key, value in data.items():
                if key == 'SprinterCommonsRegion':
                    if 'SelectedInstanceName' in value:
                        value['SelectedInstanceName'] = CC_INSTANCE_NAME
                test_log.info("{} : {}".format(key,value))
            file.seek(0)
            test_log.info("Updated values of AppConfig.yml ====>")
            yaml.dump(data, file, sort_keys=False)
            file.truncate()
        except Exception as ex:
            test_log.error("Exception occurred: {} !!! ".format(ex))
            return False

    test_log.info("Sprinter instance changed successfully and started running as a service..")
    return True


def uninstall_python_host():
    try:
        cwd = r"C:\tools"
        test_log.info("Searching for Python in Host ..")
        find_python_packages_cmd = "wmic product get name | findstr /R /C:'Python '"
        cmd_res = subprocess.run(["powershell", "-c", find_python_packages_cmd], check=True, stdout=subprocess.PIPE,
                                 stderr=subprocess.PIPE, universal_newlines=True)

        stdout, stderr = cmd_res.stdout, cmd_res.stderr
        python_pgm_list = stdout.strip().split("\n")
        stripped_python_pgm_list = [s.strip() for s in python_pgm_list]
        test_log.info(stripped_python_pgm_list)
        if cmd_res.returncode != 0:
            test_log.error("An error occurred: {}".format(stderr))
        else:
            test_log.info("Command executed successfully to fetch the python installation suite..")
        if len(stripped_python_pgm_list) != 0:
            for line in stripped_python_pgm_list:
                python_version = "python-3.8.10-amd64.exe"
                python_path = "C:\Python38"
                break
                # if str(line).find(r"Python 3.6") != -1:
                #     python_version = "python-3.6.8-amd64.exe"
                #     python_path = "C:\Python36"
                #     break
                # elif str(line).find(r"Python 3.8") != -1:
                #     python_version = "python-3.8.10-amd64.exe"
                #     python_path = "C:\Python38"
                #     break
        else:
            test_log.info("Python not found ..")
            return False

        test_log.info("Downloading {} ..".format(python_version))
        try:
            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/{} --output C:\\tools\{}".format(
                    python_version, python_version),
                shell=True)
            test_log.info("{} successfully downloaded from artifactory..".format(python_version))
        except Exception as ex:
            test_log.error("{} failed to download due to : {}".format(python_version, ex))
            return False

        uninstall_python_cmd = r"C:\tools\{} /uninstall /quiet".format(python_version)

        test_log.info("{} un-installation has started..".format(python_version))

        # installing python
        process = subprocess.run(uninstall_python_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)

        stdout, stderr = process.stdout, process.stderr

        if process.returncode == 0:
            test_log.info("{} un-installed successfully".format(python_version))
        else:
            test_log.error(stderr)
            return False
        try:
            if os.path.exists("C:\Python36"):
                shutil.rmtree("C:\Python36")
                test_log.info("Python path 'C:\Python36' has been deleted successfully..")
            elif os.path.exists("C:\Python38"):
                shutil.rmtree("C:\Python38")
                test_log.info("Python path 'C:\Python38' has been deleted successfully..")
            else:
                test_log.info("Python path not found to delete..")
        except OSError as e:
            test_log.error("Error: %s - %s." % (e.filename, e.strerror))
        return True
    except Exception as ex:
        test_log.info("Unable to find any python pre-installed..".format(ex))


def install_python():
    try:
        cwd = r"C:\tools"

        test_log.info("{} stepping has been selected.".format(str(sys_config_dict["cpu_stepping"])))
        python_version = "python-3.8.10-amd64.exe"
        python_path = "C:\Python38"

        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_version = "python-3.8.10-amd64.exe"
        #     python_path = "C:\Python38"
        # else:
        #     python_version = "python-3.6.8-amd64.exe"
        #     python_path = "C:\Python36"

        test_log.info("{} stepping has been selected.".format(str(sys_config_dict["cpu_stepping"])))

        test_log.info("Downloading {} ..".format(python_version))
        try:
            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/{} --output C:\\tools\{}".format(python_version, python_version),
                shell=True)
            test_log.info("{} successfully downloaded from artifactory..".format(python_version))

        except Exception as ex:
            test_log.error("{} failed to download due to : {}".format(python_version, ex))
            return False

        install_python_cmd = r"C:\tools\{} /quiet InstallAllUsers=1 DefaultAllUsersTargetDir={} " \
                             r"InstallLauncherAllUsers=1 Include_doc=1 Include_dev=1 PrependPath=1 Include_test=0 " \
                             r"Include_lib=1 Include_pip=1 Include_tools=1 Include_doc=1 Include_exe=1 " \
                             r"Include_tcltk=1 Include_symbols=1".format(python_version, python_path)

        test_log.info("Proceeding with fresh {} installation..".format(python_version))

        test_log.info("{} installation has started..".format(python_version))

        # installing python
        process = subprocess.run(install_python_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)

        stdout, stderr = process.stdout, process.stderr

        if process.returncode == 0:
            test_log.info("{} installed successfully".format(python_version))
        else:
            test_log.error(stderr)
            return False

        try:
            if os.path.exists("C:\Python36"):
                shutil.rmtree("C:\Python36")
            # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
            #     if os.path.exists("C:\Python36"):
            #         shutil.rmtree("C:\Python36")
            # else:
            #     if os.path.exists("C:\Python38"):
            #         shutil.rmtree("C:\Python38")
        except OSError as e:
            test_log.error("Error: %s - %s." % (e.filename, e.strerror))

        add_reg_sz_key_env_path("ITH_PYTHON_DIR", python_path + "\\")

        return True
    except Exception as ex:
        test_log.error(ex)


def install_pip():
    try:
        python_version = "python-3.8.10-amd64.exe"
        python_path = "C:\Python38"

        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_version = "python-3.8.10-amd64.exe"
        #     python_path = "C:\Python38"
        # else:
        #     python_version = "python-3.6.8-amd64.exe"
        #     python_path = "C:\Python36"

        # installing pip
        pip_install_cmd = "{}\python.exe -m ensurepip --default-pip".format(python_path)
        # installing python
        process = subprocess.Popen(pip_install_cmd, creationflags=subprocess.CREATE_NEW_CONSOLE,
                                   stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True)

        stdout, stderr = process.communicate()

        # test_log.info(process.returncode)
        if process.returncode == 0:
            test_log.info("pip installed for {} successfully".format(python_version))
            return True
        else:
            test_log.error("{} installed but pip installation failed due to : {}".format(python_version, stderr))
            time.sleep(60)
            return False

    except Exception as ex:
        test_log.error(ex)
        return False


def install_notepad_plus_plus():
    try:
        cwd = r"C:\\tools"
        test_log.info("Searching for Notepad++ installation on host ..")
        opt = subprocess.Popen(r"where /R C:\\ notepad++.exe", stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
        opt = opt.stdout.read()

        if str(opt).find(r"notepad++.exe") != -1:
            test_log.info("Notepad++ already installed ..")
            test_log.info("Notepad++ being uninstalled ..")
            output = opt.decode()
            opt = Path(output.replace("notepad++.exe\r\n", "uninstall.exe"))

            try:
                if os.path.exists(opt):
                    uninstall_notepad = "start-process -FilePath '{}' -ArgumentList '/S' -Verb runas -Wait".format(opt)

                    cmd_res = subprocess.run(["powershell", "-c", uninstall_notepad], check=True, stdout=subprocess.PIPE,
                                             stderr=subprocess.PIPE, universal_newlines=True)

                    stdout, stderr = cmd_res.stdout, cmd_res.stderr

                    test_log.info(cmd_res.returncode)
                    if cmd_res.returncode == 0:
                        test_log.info("Notepad++ uninstalled successfully")
                    else:
                        test_log.debug("Notepad++ not installed.. proceeding to install")
                else:
                    test_log.debug("Notepad++ not installed.. proceeding to install")
            except Exception as ex:
                test_log.debug("Notepad++ not installed.. proceeding to install")

        try:
            npp_path = r"C:\tools\npp.8.2.1.Installer.exe"
            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/npp.8.2.1.Installer.exe --output {}".format(npp_path),
                shell=True)
            test_log.info("Notepad++ file downloaded from artifactory successfully..")

        except Exception as ex:
            test_log.error("Notepad++ file download Failed from artifactory..")
            return False

        install_notepad_cmd = r"C:\tools\npp.8.2.1.Installer.exe /S"

        # installing notepad++
        process = subprocess.run(install_notepad_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)
        if process.returncode == 0:
            test_log.info("Notepad++ installed successfully")
        return True
    except Exception as ex:
        test_log.error(ex)


def install_ccb_package_xmlcli_windows_sut():
    sftp_client = None
    ssh = None

    cwd = r"C:\\tools"
    try:
        ccb_path = r"C:\tools\CCBSDK-Win-singleInstaller.zip"
        subprocess.check_output(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/CCBSDK-Win-singleInstaller.zip --output {}".format(ccb_path),
            shell=True)
        test_log.info("CCBSDK-Win-singleInstaller file downloaded from artifactory successfully..")

    except Exception as ex:
        test_log.error("CCBSDK-Win-singleInstaller file download Failed from artifactory..")
        return False

    local_cwd = cwd + "\CCBSDK-Win-singleInstaller.zip"
    sut_path = "C:\\CCBSDK-Win-singleInstaller.zip"
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        try:
            ssh.connect(hostname=sys_config_dict["sut_ip_address"], username=sys_config_dict["sut_login_username"],
                        password=sys_config_dict["sut_login_password"], port=22)
        except socket.gaierror:
            test_log.error("SUT Ip can't be resolved, please check the SUT ip address and start the NUC setup again..")

        sftp_client = ssh.open_sftp()
        test_log.info("Copying the CCB package from host = {} to SUT = {}".format(local_cwd, sut_path))

        sftp_client.put(local_cwd, sut_path)

        test_log.info("Extracting the CCB package on SUT..")
        cmd = "powershell Expand-Archive -Path " + sut_path + " -Force -DestinationPath " + sut_path.split("-")[0]
        stdin, stdout, stderr = ssh.exec_command(cmd)
        stdout = stdout.read()
        stderr = stderr.read().decode()
        if not stderr:
            test_log.info("Extracting the CCB package on SUT is successfull..")
        else:
            test_log.error(stderr)
            test_log.error("Extracting the CCB package on SUT Failed.. Exiting..")
            return False
        # test_log.info(stdout)
        find_exe = "cd " + sut_path.split("-")[0] + " && dir | findstr .exe"
        stdin, stdout, stderr = ssh.exec_command(find_exe)
        stdout = stdout.readlines()
        stderr = stderr.read().decode()
        if not stderr:
            test_log.info("Found the CCB package exe on SUT..")
        else:
            test_log.info(stderr)
            test_log.info("CCB package exe not available on SUT.... Exiting..")
            return False
        # test_log.info(stdout)
        exe_name = stdout[-1].strip().split("\r\n")[0].split()[-1].strip()
        test_log.info("CCB exe package name is: {}".format(exe_name))
        install_ccb_cmd_silent = "cd {} && {} -s".format(sut_path.split("-")[0], exe_name)
        test_log.info("Installing the CCB package on SUT for xmlcli support.")
        stdin, stdout, stderr = ssh.exec_command(install_ccb_cmd_silent)
        stdout = stdout.read()
        stderr = stderr.read().decode()
        if not stderr:
            test_log.info("Installation of CCB package on SUT is successful..")
        else:
            test_log.info(stderr)
            test_log.info("CCB package on SUT is not installed.. Exiting..")
            return False

        return True
    except Exception as e:
        test_log.error(e)
    finally:
        sftp_client.close()
        ssh.close()


def sut_copy_xmlcli_lin():
    sftp_client = None
    ssh = None

    cwd = r"C:\\tools"
    opt = subprocess.Popen(
        r"copy /y \\\\bdcspiec010.gar.corp.intel.com\File_Transfers\Automation_status\EGS\xmlcli.zip "
        r"{}".format(cwd),
        cwd=cwd, stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
    opt = opt.stdout.read()

    if str(opt).find(r"1 File(s) copied") != -1:
        test_log.info("XMLCLI folder downloaded successfully..")
    else:
        test_log.info(opt)
        test_log.error("XMLCLI file download Failed..")
        return False

    local_cwd_zip = cwd + r"\6.8_xmlcli_windows_linux_Python2&3_ipclean.zip"
    with zipfile.ZipFile(local_cwd_zip, 'r') as zip_ref:
        zip_ref.extractall("C:/xmlcli")
        zip_ref.close()
    local_cwd = "C:/xmlcli"
    sut_path = "opt/APP/"
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        try:
            ssh.connect(hostname=sys_config_dict["sut_ip_address"], username=sys_config_dict["sut_login_username"],
                        password=sys_config_dict["sut_login_password"], port=22)
        except socket.gaierror:
            test_log.error("SUT Ip can't be resolved, please check the SUT ip address and start the NUC setup again..")

        sftp_client = ssh.open_sftp()
        test_log.info("Copying the XMLCLI package from host = {} to SUT = {}".format(local_cwd, sut_path))

        sftp_client.put(local_cwd, sut_path)

    except Exception as e:
        test_log.error(e)
    finally:
        sftp_client.close()
        ssh.close()


def sut_xmlcli_path_check():
    ssh = None
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        try:
            ssh.connect(hostname=sys_config_dict["sut_ip_address"], username=sys_config_dict["sut_login_username"],
                        password=sys_config_dict["sut_login_password"], port=22)
        except socket.gaierror:
            # root.destroy()
            test_log.error("SUT Ip can't be resolved, please check the SUT ip address and start the NUC setup again..")

        test_log.info("Check if xmlcli is available...")

        if sys_config_dict["sut_os_name"].lower() == "rhel" or sys_config_dict["sut_os_name"].lower() == "centos" \
                or sys_config_dict["sut_os_name"].lower() == "esxi":
            cmd = r"ls /opt/APP/xmlcli/"
        else:
            cmd = r"powershell test-path -path C:\BKCPkg\xmlcli\\"

        stdin, stdout, stderr = ssh.exec_command(cmd)

        stdout = stdout.read()
        test_log.info(stdout)
        stderr = stderr.read().decode()

        if not stderr:
            test_log.info("Xmlcli directly exists ...")
        else:
            test_log.info(stderr)
            test_log.info("Xmlcli directly does not exists.. Exiting..")
            return False
        # test_log.info(stdout)
    except Exception as e:
        test_log.error("Failed to check xmlcli path due to {}".format(e))
    finally:
        ssh.close()


def sut_python_check():
    ssh = None
    try:
        ssh = paramiko.SSHClient()
        ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy)
        try:
            ssh.connect(hostname=sys_config_dict["sut_ip_address"], username=sys_config_dict["sut_login_username"],
                        password=sys_config_dict["sut_login_password"], port=22)
        except socket.gaierror:
            # root.destroy()
            test_log.error("SUT Ip can't be resolved, please check the SUT ip address and start the NUC setup again..")

        test_log.info("Check if python is available on SUT...")
        cmd = r"python --version"

        stdin, stdout, stderr = ssh.exec_command(cmd)
        stdout = stdout.read()
        test_log.info(stdout)
        stderr = stderr.read().decode()
        if not stderr:
            test_log.info("python installed on SUT...")
        else:
            test_log.info(stderr)
            test_log.error(
                "ERROR!!! --> Python does not found .. please check the SUT before triggering any tests, Exiting..")
        # test_log.info(stdout)
    except Exception as e:
        test_log.error("Failed to check Python on SUT due to {}".format(e))
    finally:
        ssh.close()

dst = r'C:\tools'


def __Run_cmd(cmd):
    try:
        test_log.info('RUN CMD >>: {0}'.format(cmd))
        ret = os.system(cmd)
        if ret != 0:
            return False
        return True
    except Exception as ex:
        test_log.error(ex)
        return False


def git_bash():
    # Use a breakpoint in the code line below to debug your script.
    # subprocess.check_output()
    try:
        test_log.info("Installing Git Bash")

        git_path = r'C:\Program Files\Git\bin'

        # force uninstall
        if os.path.exists(git_path):
            test_log.info("git already present, uninstalling git")
            g_path = r'powershell  C:\Program` Files\Git\unins*.exe /SP /VERYSILENT /NORESTART /SUPPRESSMSGBOXES ' \
                     r'/FORCECLOSEAPPLICATIONS'
            try:
                git_out = subprocess.check_output(g_path, shell=True)
            except subprocess.CalledProcessError as grepexc:
                test_log.info("git uninstalled successfully.")
                test_log.info("Return code : {} // Output : {}".format(grepexc.returncode, grepexc.output))
            except Exception as ex:
                test_log.error("git not uninstalled :{}".format(git_out))

        try:
            if not os.path.exists(git_path):
                test_log.info('Installing Git')
                try:
                    try:
                        git_bash_path = r"C:\tools\Git-2.35.1.2-64-bit.exe"
                        subprocess.check_output(
                            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/Git-2.35.1.2-64-bit.exe --output {}".format(git_bash_path),
                            shell=True)
                        test_log.info("Git-2.35.1.2-64-bit.exe file downloaded from artifactory successfully..")

                    except Exception as ex:
                        test_log.error("Git-2.35.1.2-64-bit.exe file download from artifactory Failed..")
                        return False

                    git_dst = 'powershell ' + dst + '\Git-*.exe' + ' /SP /VERYSILENT /SUPPRESSMSGBOXES ' \
                                                                   '/NORESTART /FORCECLOSEAPPLICATIONS'
                    test_log.info(git_dst)
                    __Run_cmd(git_dst)
                    test_log.info(git_path)
                    time.sleep(3)
                    if os.path.exists(git_path):
                        test_log.info('Git installation done')

                # If destination is a directory.
                except IsADirectoryError:
                    test_log.error("Destination is a directory.")

                # If there is any permission issue
                except PermissionError:
                    test_log.error("Permission denied.")
            else:
                test_log.info("Git is already present")
                return False
        except Exception as ex:
            test_log.error("Git installation Failed due to: {}".format(ex))
            return False
    except Exception as ex:
        test_log.error("Git installation Failed due to: {}".format(ex))
        return False

    path_bin = "C:\\Program Files\\Git\\bin\\"
    update_reg_sz_env_path(path_bin)

    # git work check
    process = subprocess.run(["powershell", "-c",
                              "{}git --help".format(str(path_bin).replace(" ", "` "))], check=True,
                             stdout=subprocess.PIPE, stderr=subprocess.PIPE, universal_newlines=True)

    stdout, stderr = process.stdout, process.stderr

    test_log.info(process.returncode)

    if process.returncode == 0:
        test_log.info(process.stdout)
        test_log.info("Git-Bash successfully checked and it is working.")
        test_log.info("Git-Bash installation done, please go to the main page and proceed with NUC setup..")
    else:
        test_log.info(process.stderr)
        test_log.error("Git-Bash unable to start failed, please contact the developer for debugging - Naveenraj,K..")

    test_log.info('\n+==============================================================================================+\n'
                  '+==============================================================================================+')

    install_git_button.destroy()


def uninstall_putty():
    try:
        test_log.info("Searching for Putty in Host ..")
        find_putty_packages_cmd = "wmic product get name | findstr /R /C:'PuTTY '"

        cmd_res = subprocess.run(["powershell", "-c", find_putty_packages_cmd], check=True, stdout=subprocess.PIPE,
                                 stderr=subprocess.PIPE, universal_newlines=True)

        stdout, stderr = cmd_res.stdout, cmd_res.stderr
        putty_pgm_list = stdout.strip().split("\n")
        stripped_putty_pgm_list = [s.strip() for s in putty_pgm_list]
        test_log.info(stripped_putty_pgm_list)
        if cmd_res.returncode != 0:
            test_log.error("An error occurred: {}".format(stderr))
        else:
            test_log.info("Command executed successfully to fetch the putty installation suite..")

        if len(stripped_putty_pgm_list) != 0:
            for prg in stripped_putty_pgm_list:
                uninstall_putty = 'wmic product where "{}" call uninstall ' \
                                   '/nointeractive'.format("name like '{}'").format(prg)

                # test_log.info(uninstall_putty)
                cmd_res = subprocess.run(["powershell", "-c", uninstall_putty], check=True, stdout=subprocess.PIPE,
                                         stderr=subprocess.PIPE, universal_newlines=True)

                stdout, stderr = cmd_res.stdout, cmd_res.stderr

                if cmd_res.returncode != 0 and "Method execution successful" in stdout:
                    test_log.error("An error occurred: {}".format(stderr))
                else:
                    test_log.info(stdout)
                    test_log.info("Command executed successfully to uninstall : {}".format(prg))
    except Exception as ex:
        test_log.info("Unable to find any putty pre-installed..".format(ex))


def Putty_installer():
    # Use a breakpoint in the code line below to debug your script.
    # subprocess.check_output()

    uninstall_putty()

    try:
        test_log.info("Downloading Putty Installer")

        try:
            putty_path = r"C:\tools\putty-64bit-0.76-installer.msi"
            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/putty-64bit-0.76-installer.msi --output {}".format(
                    putty_path),
                shell=True)
            test_log.info("Putty installer file downloaded from artifactory successfully..")

        except Exception as ex:
            test_log.error("Putty installer file download from artifactory Failed..")
            return False

        install_putty_cmd = r"{} /quiet".format(putty_path)

        test_log.info("Putty installation has started..")

        # installing putty
        process = subprocess.run(install_putty_cmd, stdout=PIPE, stderr=PIPE, shell=True)
        test_log.info(process.returncode)

        stdout, stderr = process.stdout, process.stderr

        if process.returncode == 0:
            test_log.info("Putty installed successfully")
        else:
            test_log.error(stderr)
            return False

    except Exception as ex:
        test_log.error("Putty installation Failed due to: {}".format(ex))
        return False


def winscp():
    # Use a breakpoint in the code line below to debug your script.
    # subprocess.check_output()
    try:
        test_log.info("Installing WinScp")

        winscp_path = r'C:\Program Files (x86)\WinSCP\WinSCP.exe'

        # force uninstall
        if os.path.exists(winscp_path):
            test_log.info("winscp already present, uninstalling winscp")
            w_path = r'cd "\program files (x86)\WinSCP\" && unins000.exe /SP /VERYSILENT /SUPPRESSMSGBOXES ' \
                     r'/FORCECLOSEAPPLICATIONS /NORESTART'
            w_out = subprocess.check_output(w_path, shell=True)
            test_log.info("wscp : {}".format(w_out))

        try:
            if not os.path.exists(winscp_path):
                test_log.info('Installing winscp')
                try:
                    try:
                        winscp_path = r"C:\tools\WinSCP-5.19.5-Setup.exe"
                        subprocess.check_output(
                            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/WinSCP-5.19.5-Setup.exe --output {}".format(
                                winscp_path), shell=True)
                        test_log.info("WinSCP-5.19.5-Setup.exe file downloaded from artifactory successfully..")
                    except Exception as ex:
                        test_log.error("WinSCP-5.19.5-Setup.exe file download from artifactory "
                                       "Failed due to: {}".format(ex))
                        return False

                    winscp_dst = 'powershell ' + dst + '\WinSCP*.exe' + ' /SP- /VERYSILENT ' \
                                                                        '/SUPPRESSMSGBOXES ' \
                                                                        '/FORCECLOSEAPPLICATIONS /NORESTART'
                    test_log.info(winscp_dst)
                    __Run_cmd(winscp_dst)
                    test_log.info(winscp_path)
                    time.sleep(3)
                    if os.path.exists(winscp_path):
                        test_log.info('winscp installation done')
                    else:
                        test_log.info('path not found')

                # If destination is a directory.
                except IsADirectoryError:
                    test_log.error("Destination is a directory.")

                # If there is any permission issue
                except PermissionError:
                    test_log.error("Permission denied.")

            else:
                test_log.info("winscp is already present")
                return False

        except Exception as ex:
            test_log.error("WinScp installation Failed due to: {}".format(ex))
            return False
    except Exception as ex:
        test_log.error("WinScp installation Failed due to: {}".format(ex))
        return False


def pycharm_installer():
    import re
    # Use a breakpoint in the code line below to debug your script.
    path_found = False
    try:
        try:
            # check if Pycharm folder exist
            test_log.info("checking for existing pycharm")
            path_res = subprocess.Popen("where /R C:\\ pycharm*.exe",
                                        stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
            path_res = path_res.stdout.read()
            path_out = path_res.decode().split("\r\n")
            path_found = None
            for path in path_out:
                if os.path.exists(path):
                    if bool(re.search("pycharm[\d].*\.exe", path)):
                        test_log.info("pycharm installation folder found under: {} ".format(path))
                        path_found = True
                        break
                    else:
                        test_log.debug("pycharm installation folder not found: {} ".format(path))
        except Exception as ex:
            test_log.debug("Pycharm not found..")

        # force uninstall
        try:
            if path_found:
                path_tail = os.path.split(Path(path))[-1]
                path = path.replace(path_tail, "Uninstall.exe")
                test_log.info("pycharm already present, uninstalling pycharm..")
                test_log.info("Uninstallation path with space :{}".format(path))
                add_accent_character = str(path).replace(" ", "` ")
                test_log.info("Uninstallation path with grave accent character :{}".format(add_accent_character))
                try:
                    if os.path.exists(path):
                        uninstall_path = r'{} /S /VERYSILENT /SUPPRESSMSGBOXES ' \
                                         r'/FORCECLOSEAPPLICATIONS'.format(add_accent_character)

                        cmd_res = subprocess.run(["powershell", "-c", uninstall_path], check=True,
                                                 stdout=subprocess.PIPE,
                                                 stderr=subprocess.PIPE, universal_newlines=True)

                        stdout, stderr = cmd_res.stdout, cmd_res.stderr

                        test_log.info(cmd_res.returncode)
                        if cmd_res.returncode == 0:
                            test_log.info(stdout)
                            test_log.info("pycharm uninstalled successfully")
                        else:
                            test_log.debug(stderr)
                            test_log.debug("pycharm not installed.. proceeding to install")
                    else:
                        test_log.debug("pycharm not installed.. proceeding to install")
                except Exception as ex:
                    test_log.debug(ex)
                    test_log.debug("pycharm not installed.. proceeding to install")
                time.sleep(5)
        except Exception as ex:
            test_log.error("Pycharm uninstallation failed due to :{}".format(ex))
            return False

        try:
            try:
                test_log.info("Downloading Pycharm from artifactory")
                pycharm_exe_path = r"C:\tools\pycharm-community-2019.3.5.exe"
                pycharm_silent_file = r"C:\tools\silent.config"
                subprocess.check_output(
                    "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/pycharm-community-2019.3.5.exe --output {}".format(
                        pycharm_exe_path), shell=True)
                subprocess.check_output(
                    "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/silent.config --output {}".format(
                        pycharm_silent_file), shell=True)
                test_log.info("pycharm-community-2019.3.5.exe file downloaded from artifactory successfully..")
            except Exception as ex:
                test_log.error("pycharm-community-2019.3.5.exe file download from artifactory "
                               "failed due to: {}".format(ex))
                return False

            pycharm_dst = r'C:\tools\pycharm-community-2019.3.5.exe /S /CONFIG=c:\tools\silent.config'

            test_log.info("Installing Pycharm...")
            __Run_cmd(pycharm_dst)
            # test_log.info(pycharm_path)
            time.sleep(5)

            try:
                # check if Pycharm folder exist
                test_log.info("checking for newly Installed pycharm")
                path_res = subprocess.Popen("where /R C:\\ pycharm*.exe",
                                            stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
                path_res = path_res.stdout.read()

                path_out = path_res.decode().split("\r\n")
                for path in path_out:
                    if os.path.exists(path):
                        if bool(re.search("pycharm[\d].*\.exe", path)):
                            test_log.info("pycharm installation folder found under: {} ".format(path))
                            test_log.info('pycharm installation done')
                            return True
                        else:
                            test_log.debug("pycharm installation folder not found: {} ".format(path))
            except Exception as ex:
                test_log.debug("Pycharm not found..")

        # If destination is a directory.
        except IsADirectoryError:
            test_log.error("Destination is a directory.")
            return False
        # If there is any permission issue
        except PermissionError:
            test_log.error("Permission denied.")
            return False
        except Exception as ex:
            test_log.error("Pycharm installation Failed due to: {}".format(ex))
            return False
    except Exception as ex:
        test_log.error("Pycharm installation Failed due to: {}".format(ex))
        return False


def cscript_installer():
    try:
        test_log.info("Installing CScripts")
        test_log.info("cscript_installer started")
        c_dst = r'C:\\'

        if not os.path.exists(r"C:\Intel\DFPython\bin"):
            test_log.info("DFPython directory does not exists.. Copy the directory and proceed further")
            return False
        else:
            test_log.info("DFPython directory exists.. proceeding further")

        # dflaunch_path = "C:\\Intel\\DFPython\\bin\\"
        # update_reg_sz_env_path(dflaunch_path)

        cscript_file = r'C:\cscripts'
        if os.path.isdir(cscript_file):
            test_log.info("cscripts is already present")
            # force uninstall cscript
            import shutil
            shutil.rmtree(cscript_file)
            test_log.info("cscripts removed successfully")
        else:
            test_log.info('cscripts not found, installing Cscripts')

        try:
            cscripts_zip_path = r"C:\tools\cscripts.zip"
            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/cscripts.zip --output {}".format(
                    cscripts_zip_path), shell=True)
            test_log.info("cscripts.zip file downloaded from artifactory successfully..")
        except Exception as ex:
            test_log.error("cscripts.zip file download from artifactory Failed due to: {}".format(ex))
            return False

        with zipfile.ZipFile(cscripts_zip_path, 'r') as zip_ref:
            zip_ref.extractall(c_dst)
        test_log.info("cscripts unzipped successfully.")
        test_log.info("cscript_installer end")
        activate_virtual_env()
        c_install_pythonlib()
    except Exception as ex:
        test_log.error("cscripts installation Failed due to: {}".format(ex))
        return False


# acrivate virtual environment
def activate_virtual_env():
    try:
        version = "--version 3.8"
        active_env = "white_py38_cscript"
        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     version = "--version 3.8"
        #     active_env = "white_py38_cscript"
        # else:
        #     version = "--version 3.6"
        #     active_env = "white_py36_cscript"

        test_log.info("activate_virtual_env started")
        test_log.info("get list of venv")

        # get list of venv
        list_cmd = 'powershell cd C:\\cscripts && C:\Intel\DFPython\\bin\dflaunch.bat --list'
        dflaunch_out = subprocess.check_output(list_cmd,  shell=True)

        dflaunch_out = (str(dflaunch_out.rstrip()).strip('\\r')).split('\\n')
        test_log.info(dflaunch_out)

        current_active_env = dflaunch_out[-1].split(":")[-1].strip().replace("\'","")
        test_log.info("Current Active Environment is: {}".format(current_active_env))
        i = 0
        if current_active_env != 'None':
            test_log.info("{} present, Deleting venv".format(active_env))
            # Del exist session
            d_cmd = 'powershell cd C:\\cscripts && C:\Intel\DFPython\\bin\dfmanager delete {}'.format(current_active_env)
            dfmanager_d = subprocess.check_output(d_cmd,  shell=True)
            test_log.info((dfmanager_d))
        test_log.info("activate_virtual_env end")

        dflaunch_out = subprocess.check_output(list_cmd, shell=True)

        dflaunch_out = (str(dflaunch_out.rstrip()).strip('\\r')).split('\\n')
        test_log.info(dflaunch_out)

        for venv in dflaunch_out:
            if "white" in venv or "red" in venv:
                available_venv = str(venv).strip().replace(r"\r", "")

                # Del exist session
                d_cmd = 'powershell cd C:\\cscripts && C:\Intel\DFPython\\bin\dfmanager delete {}'.format(
                    available_venv)
                dfmanager_d = subprocess.check_output(d_cmd, shell=True)
                test_log.info(dfmanager_d)
                test_log.info("Deleted the Virtual Env :: {}".format(venv))

        # create virtual env
        test_log.info("create virtual environment starts..")
        c_cmd = 'powershell cd C:\\cscripts && C:\Intel\DFPython\\bin\dfmanager create cscripts --color white {}'.format(version)
        dfmanager_out = subprocess.check_output(c_cmd,  shell=True)
        test_log.info('Virtual env : cscripts created ...waiting to terminate....')
        dfmanager_out = (str(dfmanager_out.rstrip()).strip('\\r')).split('\\n')
        test_log.info((dfmanager_out))
        i = 0
        for i in range(len(dfmanager_out)):
            if 'Name:' in str(dfmanager_out[i]):
                cname = ((str(str(dfmanager_out[i])).strip('\\r')).split(":"))[1]
                test_log.info(cname)
        test_log.info("dfmanager create cscripts --color white {} done".format(version))
        test_log.info("create virtual done")

        # dfmanager activation
        test_log.info("dfmanager activation started")
        ac_cmd = 'powershell cd C:\\cscripts && C:\Intel\DFPython\\bin\dfmanager activate ' + str(cname)
        dfmanager_ac = subprocess.check_output(ac_cmd,  shell=True)
        dfmanager_ac = (str(dfmanager_ac.rstrip()).strip('\r')).split('\n')
        i = 0
        for i in range(len(dfmanager_ac)):
            test_log.info(str(dfmanager_ac[i]))
            if 'Environment successfully activated!' in str(dfmanager_ac[i]):
                cname = str(dfmanager_ac[i])
                log_out = 'cscripts activation done for: ' + str(cname)
                test_log.info(log_out)
        test_log.info("dfmanager 2")
        test_log.info("dfmanager activation done")
    except Exception as ex:
        test_log.error("cscripts activation Failed due to: {}".format(ex))
        return False


def c_install_pythonlib():
    try:
        python_version = "python-3.8.10-amd64.exe"
        python_path = "C:\\Python38"
        python_path_scripts = "C:\\Python38\\Scripts"
        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_version = "python-3.8.10-amd64.exe"
        #     python_path = "C:\\Python38"
        #     python_path_scripts = "C:\\Python38\\Scripts"
        # else:
        #     python_version = "python-3.6.8-amd64.exe"
        #     python_path = "C:\\Python36"
        #     python_path_scripts = "C:\\Python36\\Scripts"
        # copy installer to host
        test_log.info("cscripts pythonlib installation started")
        cscript_installer_dest = 'C:\cscripts\\Cscript_Requirement.txt'

        if not os.path.exists(cscript_installer_dest):
            test_log.info('cscript_pkg_installer not found, installing cscript_pkg_installer')
            try:
                subprocess.check_output(
                    "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/Cscript_Requirement.txt --output {}".format(
                        cscript_installer_dest), shell=True)
                test_log.info("Cscript_Requirement.txt file downloaded from artifactory successfully..")
            except Exception as ex:
                test_log.error("Cscript_Requirement.txt file download from artifactory Failed due to: {}".format(ex))
                return False
        else:
            test_log.info('cscript_pkg_installer found')

        test_log.info("cscripts pythonlib installation started with cscript_req in dflaunch")
        #Faceless account
        username = os.environ["sys_crse_user"]
        u_password = os.environ["sys_crse_pwd"]

        if str(sys_config_dict["is_this_rasp_system"]).lower() == "flex lab":
            df_l_cmd1 = 'powershell cd C:\\cscripts && dflaunch -m pip install --upgrade pip -i https://' + str(
                username) + ':' + str(u_password) + '@intelpypi.intel.com/pythonsv-white/production'
            df_l_cmd = 'powershell cd C:\\cscripts && dflaunch -m pip install -r C:\cscripts\\Cscript_Requirement.txt -i https://' + str(
                username) + ':' + str(u_password) + '@intelpypi.intel.com/pythonsv-white/production'
        else:
            df_l_cmd1 = 'powershell cd C:\\cscripts && dflaunch -m pip install --proxy=proxy-chain.intel.com:911 --upgrade pip -i https://'+str(username)+':'+str(u_password)+'@intelpypi.intel.com/pythonsv-white/production'
            df_l_cmd = 'powershell cd C:\\cscripts && dflaunch -m pip install --proxy=proxy-chain.intel.com:911 -r C:\cscripts\\Cscript_Requirement.txt -i https://' + str(
                username) + ':' + str(u_password) + '@intelpypi.intel.com/pythonsv-white/production'

        dfl_out = os.system(df_l_cmd1)
        test_log.info(dfl_out)

        dfl_out = os.system(df_l_cmd)
        p_log = "dfl out>>>" + str(dfl_out)
        test_log.info(p_log)

        # to save install logs
        test_log.info("save python install logs in >>> C:\cscripts\\Cscript_Requirement_logs.txt")
        from datetime import datetime
        currentDateTime = '>>>>>>>' + str(datetime.now()) + '\n'
        test_log.info(currentDateTime)
        log_path = "C:\cscripts\\Cscript_Requirement_logs.txt"
        with io.open(log_path, "w") as file1:
            file1.writelines(str(currentDateTime))
            file1.writelines(str(dfl_out))
        test_log.info("cscripts pythonlib installation done")
    except Exception as ex:
        test_log.error(ex)
        return False


def redfish_checkup():
    test_log.info("Redfish check started")
    try:
        if sys.platform != 'win32':
            tree = ET.parse('/home/Automation/system_configuration.xml')
        else:
            tree = ET.parse('C:\Automation\system_configuration.xml')
        root = tree.getroot()

        for item in root.iter('physical_control'):
            type = item.find('./driver/redfish/bmc_type')
            username = item.find('./driver/redfish/username')
            password = item.find('./driver/redfish/password')
            ip = item.find('./driver/redfish/ip')
        test_log.info("Extracted INFORMATION  ===> {} {} {} {}".format(ip.text, username.text, password.text, type.text))
    except Exception as ex:
        test_log.error(ex)

    try:
        ret = subprocess.check_output("ping -n 3 " + ip.text, shell=True)
        if "0% loss" in str(ret):
            test_log.info('Redfish BMC is alive')
    except Exception as ex:
        test_log.error("Platform is in Turned OFF State or Mentioned BMC IP is not Correct Ping "
                       "Failed to BMC {0}".format(ip.text))
        return False
    try:
        session = requests.session()
        requests.urllib3.disable_warnings()
        try:
            url = (
                r'https://{0}/redfish/v1/Chassis/AC_Baseboard/Power#/Voltages/0'.format(ip.text))
            auth1 = (username.text, password.text)
            reps = session.get(url, auth=auth1, verify=False)
        except Exception as ex:
            test_log.error("Redfish AC_baseboard is Different on this Platform Trying Alternative Approach")
            url = (
                r'https://{0}/redfish/v1/Chassis/BC_Baseboard/Power#/Voltages/0'.format(ip.text))
            auth1 = (username.text, password.text)
            reps = session.get(url, auth=auth1, verify=False)
        version = reps.json()
        # test_log.info(json.dumps(version, indent=4, sort_keys=True))
        for volt in version['Voltages']:
            if volt['Name'] == "P3V3":
                fp_volt = volt['ReadingVolts']
                if not fp_volt:
                    fp_volt = 0
            if volt['Name'] == "P3VBAT":
                dc_volt = volt['ReadingVolts']
                break
        test_log.info("P3V3 {0}  P3VBAT {1}".format(fp_volt, dc_volt))
        version = [fp_volt, dc_volt]
        if "3." in str(version):
            test_log.info('RedFish BMC is UP ------------------------------- [OK]')
            return True
        else:
            test_log.info("Redfish BMC Credentials are Wrong ------------------------------ [BAD]")
    except Exception as ex:
        test_log.error("Failed to Get Volt Value >> {0}".format(ex))


def execute(list1):
    try:
        # useing comma "," not space " " in Popen[]. replace the key word "*" to what you want to search.
        output = subprocess.Popen(list1, stdout=subprocess.PIPE,
                                  stderr=subprocess.PIPE)
        stdout, stderr = output.communicate()
    except subprocess.CalledProcessError:
        test_log.error('Exception handled')
    except Exception as ex:
        test_log.error("stderr: \n".format(stderr))
        test_log.error('Exception occurred due to : {}'.format(ex))
        return False
    return stdout


def itp_probe_check_device_manager():
    try:
        execute(
            "curl -X GET https://ubit-artifactory-ba.intel.com/artifactory/list/dcg-dea-srvplat-repos/Automation_Tools/"
            "SPR/Windows/devcon.exe --output devcon.exe")
        out = execute(["devcon.exe", "hwids", "=ITP_ON_USB"])
        l1 = out.split(b"USB\\Class_FF\r\n")
        if l1 != None or l1[1][0] != 48:
            test_log.info("Found ITP probe !!!")
        else:
            test_log.info("ITP probe not found !!!")
    except Exception as ex:
        test_log.error(ex)


def pythonsv_installer_semi_auto():
    try:
        python_path = "C:\\Python38"
        python_path_scripts = "C:\\Python38\\Scripts"
        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_path = "C:\\Python38"
        #     python_path_scripts = "C:\\Python38\\Scripts"
        # else:
        #     python_path = "C:\\Python36"
        #     python_path_scripts = "C:\\Python36\\Scripts"

        installer_path_zip = "c:\\tools\\InstallerPySV.zip"
        subprocess.check_output(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/"
            "dcg-dea-srvplat-local/Automation_Tools/IPU/ipu_infra/InstallerPySV.zip --output {}".format(
                installer_path_zip), shell=True)
        test_log.info("InstallerPySV.zip file copied under c:\\tools directory successfully...")
        with zipfile.ZipFile(installer_path_zip, 'r') as zip_ref:
            zip_ref.extractall("c:\\tools")

        test_log.info("InstallerPySV.zip file extracted under c:\\tools directory successfully...")

        path_res = subprocess.Popen("where /R C:\\ InstallMe.py",
                                    stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)

        path_res = path_res.stdout.read()

        if str(path_res).find(r"InstallMe") != -1:
            test_log.info("InstallMe.py file not found under {}: ".format(path_res))
        else:
            test_log.debug("InstallMe.py file found {}: ".format(path_res))

        path_out = path_res.decode().split("\r\n")
        path_found = None
        for path in path_out:
            if os.path.exists(path):
                if "InstallMe.py" in path:
                    test_log.info("InstallMe.py file found under: {} ".format(path))
                    path_found = True
                    break
                else:
                    test_log.debug("InstallMe.py file not found: {} ".format(path))
                    path_found = False

        if path_found:
            try:
                run_pip = "{}\pip3 install requests --proxy=http://proxy01.iind.intel.com:911".format(python_path_scripts)
                pip_install_pac = subprocess.Popen(run_pip, creationflags=subprocess.CREATE_NEW_CONSOLE, stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)

                pip_install_pac = pip_install_pac.stdout.read()

                test_log.info(pip_install_pac)

                if str(pip_install_pac).find(r"already satisfied: requests") != -1 or \
                        str(pip_install_pac).find(r"Successfully installed") != -1:
                    test_log.info("requests module installed")
                else:
                    test_log.error("requests module did not install.. trying with pip")

                    run_pip = "{}\pip install requests --proxy=http://proxy01.iind.intel.com:911".format(python_path_scripts)
                    pip_install_pac = subprocess.Popen(run_pip, creationflags=subprocess.CREATE_NEW_CONSOLE, stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)

                    pip_install_pac = pip_install_pac.stdout.read()

                    test_log.info(pip_install_pac)

                    if str(pip_install_pac).find(r"already satisfied: requests") != -1 or \
                            str(pip_install_pac).find(r"Successfully installed") != -1:
                        test_log.info("requests module installed")
                    else:
                        test_log.error("Pythonsv not configured due to request module not found...")
                        test_log.error("requests module did not install.. Exiting..")
                        return False
            except Exception as ex:
                test_log.error(ex)

            test_log.info("Pythonsv semi auto installation started, please configure when the popup comes up..")

            path_res = subprocess.Popen("{}\python {}".format(python_path, path), stdin=PIPE, stdout=PIPE,
                                        stderr=STDOUT, shell=True)

            stdout, stderr = path_res.communicate()

            if stderr is None:
                test_log.info("Hope you have configured the Pythonsv correctly..")
    except Exception as ex:
        test_log.error(ex)


def python_check():
    try:
        test_log.info('Python Version Check has started...')
        q = subprocess.check_output('python -V', shell=True)
        q = str(q)
        if "3" in q:
            test_log.info("Python 3 is installed . Version that is installed is {}: ".format(q))
            return True
        else:
            test_log.debug("Python 3 is not installed . Output of python -Version is : ".format(q))
            test_log.debug("#################### [ Python 3 is not installed ] ########################")
            return False
    except Exception as ex:
        test_log.debug(ex)
        return False


def python_lib_check(filename):
    try:
        python_path = "C:\\Python38"
        python_path_scripts = "C:\\Python38\\Scripts"
        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_path = "C:\\Python38"
        #     python_path_scripts = "C:\\Python38\\Scripts"
        # else:
        #     python_path = "C:\\Python36"
        #     python_path_scripts = "C:\\Python36\\Scripts"

        clearlines = []
        installed = []
        noninstalled = []
        flag = 0
        test_log.info('Python Library check has started...')
        test_log.info("Checking if all requirements of : '{}' : are present".format(filename))
        fileobject = io.open(filename)
        filetext = fileobject.read()
        filelines = filetext.split('\n')
        for a in filelines:
            if (a.startswith('#')):
                continue
            else:
                if (a.startswith('--')):
                    a = a.split('=')
                    a = a[1]
                    if "," in a:
                        a = a.split(',')
                        for i in a:
                            clearlines.append(i)
                else:
                    if (a == ''):
                        continue
                    else:
                        if "=" in a:
                            a = a.split("=")
                            a = a[0]
                            if "<" in a:
                                a = a.split('<')
                                a = a[0]
                                clearlines.append(a)
                                continue
                            if ">" in a:
                                a = a.split('>')
                                a = a[0]
                                clearlines.append(a)
                                continue
                            clearlines.append(a)
                        else:
                            if "<" in a:
                                a = a.split('<')
                                a = a[0]
                                clearlines.append(a)
                                continue
                            if ">" in a:
                                a = a.split('>')
                                a = a[0]
                                clearlines.append(a)
                                continue
                            clearlines.append(a)

        test_log.info('List of libraries that are present in the files are : ')
        test_log.info(clearlines)
        test_log.info('Running "pip list" to get list of all installed libraries')
        q = subprocess.check_output('{}\pip list'.format(python_path_scripts), creationflags=subprocess.CREATE_NEW_CONSOLE, shell=True)
        q = str(q)
        q = q.lower()
        for a in clearlines:
            a = str(a)
            a = a.replace('_', '-')
            a = a.lower()
            if q.__contains__(str(a)):
                installed.append(a)
            else:
                noninstalled.append(a)
                flag = 1
        test_log.info('Comparing the requirement file to the pip list')
        if (flag == 0):
            test_log.info("All packages mentioned in '{}' in the requirement file are found".format(filename))
        else:
            test_log.debug('List of packages that are not found are :')
            test_log.debug(noninstalled)
            test_log.debug("Some packages are not found.")

        fileobject.close()
    except Exception as ex:
        test_log.error(ex)


def quartus_uninstall():
    test_log.info("Starting to check quartus application for uninstalling ..")

    opt = subprocess.Popen(r"where /R C:\\ quartus_pgm.exe", stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
    opt = opt.stdout.read()
    path_out = opt.decode().split("\r\n")
    uninstall_res = 0
    for opt_p in path_out:
        if str(opt_p).find(r"quartus_pgm.exe") != -1:
            test_log.info(opt_p)
            test_log.info("quartus application is already installed ..")
            cmd = os.path.split(opt_p)[0].replace(r"qprogrammer\bin64", "uninstall")
            test_log.info(cmd)
            test_log.info("quartus application is found ..")

            opt = subprocess.Popen(r"where /R {} *.exe".format(cmd), stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
            opt = opt.stdout.read()
            cmd_uninstall = opt.decode().strip().split("\r\n")[0]

            test_log.info("Starting the un-installer of quartus application ...")
            test_log.info(cmd_uninstall)
            # C:\intelFPGA_pro\18.1\uninstall\qprogrammer-18.1.0.222-windows-uninstall.exe
            opt_quartus_uninstall = subprocess.Popen(r"{} --mode unattended".format(cmd_uninstall),
                stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)

            opt_quartus_uninstall.wait()
            if opt_quartus_uninstall.returncode == 0:
                test_log.info("quartus application uninstalled successfully..")
            else:
                test_log.error("Failed to uninstall quartus application..")
        else:
            uninstall_res += 1

    if len(path_out) == uninstall_res:
        test_log.info("quartus application is not installed.")


def quartus_checkup():

    cmd = None
    test_log.info("Quartus Checkup Started")
    if os.path.exists("C:\Automation\system_configuration.xml"):
        if sys.platform != 'win32':
            tree = ET.fromstring('/home/Automation/system_configuration.xml')
        else:
            tree = ET.parse('C:\Automation\system_configuration.xml')
        try:
            root = tree.getroot()
            for item in root.findall('./suts/sut/providers'):
                cpld_app_path = item.find('./flash/driver/usbblaster/cpld_application_path')
                test_log.info(cpld_app_path.text)
                if os.path.exists(str(cpld_app_path.text)) and "18.1" in str(cpld_app_path.text):
                    test_log.info("PATH Exists for Quartus Application")
                    cmd = cpld_app_path.text
                else:
                    test_log.info("PATH For Quartus Application Doesn't Exists Check path in system_configuration "
                                  "File or Install quartus Application")
        except Exception as ex:
            test_log.debug("Mismatched Syntax in system_configuration.xml Properly Check the syntax {0}".format(ex))
            test_log.info("Creating new System configuration file...")
    else:
        test_log.debug("System configuration file is yet to be generated..")

    test_log.info("Cross checking the NUC one more time for the presence of Quartus 18.1 application")

    opt = subprocess.Popen(r"where /R C:\\ quartus_pgm.exe", stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)
    opt = opt.stdout.read()
    path_out = opt.decode().split("\r\n")

    if cmd is None:
        for opt_p in path_out:
            if str(opt_p).find(r"quartus_pgm.exe") != -1 and str(opt_p).find(r"18.1") != -1:
                test_log.info(opt_p)
                test_log.info("quartus_pgm 18.1 is already installed ..")
                cmd = opt_p
                break

    path_out = opt
    if cmd is not None and "18.1" in cmd:
        test_log.info(cmd)
        test_log.info("quartus_pgm 18.1 is found ..")
    else:
        test_log.info("quartus_pgm 18.1 is not installed previously..")
        test_log.info("quartus_pgm 18.1 is downloading from artifactory..")
        subprocess.check_output(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local/Automation_Tools/SPR/Auto_Installer_Kit/collaterals/QuartusProProgrammerSetup-18.1.0.222-windows.exe --output C:\\tools\QuartusProProgrammerSetup-18.1.0.222-windows.exe",
            shell=True)
        test_log.info("quartus_pgm 18.1 downloaded from artifactory..")
        test_log.info("Starting the installer of quartus_pgm 18.1 ...")
        opt_quartus = subprocess.Popen(
            r"C:\\tools\QuartusProProgrammerSetup-18.1.0.222-windows.exe --unattendedmodeui none --mode unattended --accept_eula 1",
            stdin=PIPE, stdout=PIPE, stderr=STDOUT, shell=True)

        opt_quartus.wait()
        if opt_quartus.returncode == 0:
            test_log.info("quartus_pgm 18.1 installed successfully..")
            test_log.info("Searching for quartus_pgm 18.1 application..")
            opt = subprocess.Popen(r"where /R C:\\ quartus_pgm.exe", stdin=PIPE, stdout=PIPE, stderr=STDOUT,
                                   shell=True)
            opt = opt.stdout.read()
            path_out = opt.decode().split("\r\n")
            test_log.info(path_out)
            cmd = path_out[0]
            path_out = path_out[0]
        else:
            test_log.debug("Failed to install quartus_pgm 18.1..")

        test_log.info("Cross checking possible installation path of quartus 18.1 application..")

    possible_quartus_path = [r"C:\intelFPGA\18.1\qprogrammer\bin64\quartus_pgm.exe",
                             r"C:\intelFPGA_pro\18.1\qprogrammer\bin64\quartus_pgm.exe"]

    if os.path.exists(possible_quartus_path[0]):
        if cmd is None : cmd = possible_quartus_path[0]
        final_quart_path_config = os.path.split(possible_quartus_path[0])[0]
    elif os.path.exists(possible_quartus_path[1]):
        if cmd is None : cmd = possible_quartus_path[1]
        final_quart_path_config = os.path.split(possible_quartus_path[1])[0]
    else:
        test_log.error("quartus_pgm 18.1 did not install ..")
        return False

    test_log.info("Make Sure SUT is Turned ON for Detecting CPLD...")

    if cmd is not None:
        test_log.info("Quartus App path from system configuration is not found, found path after installing quartus.."
                      "proceeding further..")
        test_log.info(cmd)
        quartus_split_path = os.path.split(cmd)[0]
        cmd = "{} -c 1 -a ".format(cmd)
    elif path_out is not None:
        test_log.info("Quartus App path from system configuration is not found, while cross checking found the quartus path... "
                      "proceeding further..")
        test_log.info(path_out)
        quartus_split_path = os.path.split(path_out)[0]
        cmd = "{} -c 1 -a ".format(path_out)
    else:
        test_log.info("Quartus App path from system configuration is available, proceeding further..")
        test_log.info(cmd)
        quartus_split_path = os.path.split(cmd)[0]
        cmd += " -c 1 -a "

    try:
        ret = subprocess.check_output(cmd, shell=True)
        # print("{0}".format(ret))  #debug
        if "Quartus Prime Programmer was successful. 0 errors, 0 warnings" in str(ret):
            test_log.info("Quartus Programming USB Blaster is Detected In Host Machine ----------------------- [OK]")
    except Exception as ex:
        test_log.error("Errors Quartus USB Blaster Is Not Detected {0} -------------- [FAILED]".format(ex))
    rpath = pathlib.Path("C:\\postcode\\rfat_modified.tcl")
    try:
        test_log.info("Checking whether we have c:\\postcode directory...")
        subprocess.check_output("rd /s /q c:\\postcode", shell=True)
        test_log.info("c:\\postcode directory is present...")
    except Exception as ex:
        test_log.debug("c:\\postcode directory is not present...{}".format(ex))

    if rpath.exists() == False:
        test_log.info("Creating c:\\postcode directory ...")
        subprocess.check_output("mkdir C:\\postcode", shell=True)
        test_log.info("Copying postcode.zip under c:\\postcode directory ...")
        time.sleep(3)
        result = subprocess.Popen(
            "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/dcg-dea-srvplat-local"
            "/Automation_Tools/CI_Tools/postcode.zip --output c:\\postcode\\postcode.zip", stdout=PIPE, stderr=PIPE)
        stdout, stderr = result.communicate()
        test_log.debug("Output {}".format(stdout))
        test_log.debug("STDERR of the download command : {}".format(stderr))
        if str(stdout).find(r"100") != -1 or str(stderr).find(r"100") != -1:
            test_log.info("Command ran successfully")
        else:
            test_log.error("Failed to run the copy command")
        test_log.info("postcode.zip file copied under c:\\postcode directory successfully...")
        with zipfile.ZipFile("C:\\postcode\\postcode.zip", 'r') as zip_ref:
            zip_ref.extractall("c:\\postcode")

        test_log.info("postcode.zip file extracted under c:\\postcode directory successfully...")
    else:
        test_log.info("rfat tcl File For Reading Postcode Already Exists")

    try:
        if isinstance(quartus_split_path, str):
            ecmd = "cd C:\\postcode && " + quartus_split_path + "\quartus_stp_tcl -t rfat_modified.tcl"
            test_log.info("The path is {}".format(ecmd))
        elif isinstance(quartus_split_path, tuple):
            ecmd = "cd C:\\postcode && " + quartus_split_path[0] + "\quartus_stp_tcl -t rfat_modified.tcl"
            test_log.info("The path is {}".format(ecmd))
        else:
            test_log.debug("The path is {}".format(quartus_split_path))
            test_log.error("Unable to determine the correct path for running 'quartus_stp_tcl' with quartus, "
                           "try again or contact the developer for debugging, Naveenraj K")
            return False

        if ecmd is None:
            test_log.error("Unable to determine the correct path for running 'quartus_stp_tcl' with quartus, "
                           "try again or contact the developer for debugging, Naveenraj K")
            return False

        # print(ecmd) #debug
        output = Popen(ecmd, shell=True, stdin=PIPE, stdout=PIPE, stderr=STDOUT)
        cmd = output.stdout.read()
        cmd = str(cmd)
        ind = (str(cmd).index("BIOS CODE:"))
        ind1 = (str(cmd).index("FPGA CODE:"))
        psd = []
        post_code = (cmd[int(ind + 11):int(ind + 13)])
        fpga_code = (cmd[int(ind1 + 11):int(ind1 + 13)])
        psd = post_code, fpga_code
        test_log.info("Platform FPGA Code {0} Platform BIOS Code {1} ------------------------ [OK]".format(psd[0],
                                                                                                              psd[1]))
        return final_quart_path_config
    except Exception as ex:
        test_log.error("Errors Caught While Reading Postcode {0}".format(ex))
    finally:
        return final_quart_path_config

def banino_sx_sft_prepare():
    test_log.info("Make sure Banino Controller and Banino folder is closed..")

    try:
        if not os.path.isfile("c:\\banino"):
            try:
                shutil.rmtree("c:\\banino")
            except Exception as ex:
                test_log.debug(ex)
            try:
                os.mkdir("c:\\banino")
            except Exception as ex:
                test_log.debug(ex)

            subprocess.check_output(
                "curl -u "+ str(atf_username) + ":" + str(atf_password) + " -X GET https://ubit-artifactory-ba.intel.com/artifactory/"
                "dcg-dea-srvplat-local/Automation_Tools/CI_Tools/banino.zip --output C:\\tools\\banino.zip",
                shell=True)
            time.sleep(5)
            with zipfile.ZipFile("c:\\tools\\banino.zip", 'r') as zip_ref:
                zip_ref.extractall("c:\\banino\\")
                zip_ref.close()
            test_log.info("Banino files Extraction To C: drive Successful..")
            test_log.info("Download LadyBird flash, Banino Application, SX code ------------------------------- [OK]")
        else:
            test_log.info("Banino zip folder is already present.. ")
        return True
    except Exception as ex:
        test_log.error(ex)


def banino_serial_setup():
    ladybird = ctypes.cdll.LoadLibrary(r"C:\banino\code\Banino_SXState\x64\ladybird.dll")
    buffer = (ctypes.c_int * 32)()
    lbBuffer = (ctypes.c_byte * 512)()
    ladybirddDevices = ladybird.findDevices(512, ctypes.c_void_p(ctypes.addressof(lbBuffer)))
    for i in range(2):
        s = str(hex(
            (lbBuffer[i * 12 + 7] << 24) + (lbBuffer[i * 12 + 6] << 16) + (lbBuffer[i * 12 + 5] << 8) + lbBuffer[
                i * 12 + 4]))
    res = int(s, 16)
    lb_serial = (str(res))
    # print(lb_serial)
    file_name = r'C:\banino\code\Banino_SXState\config.xml'
    with fileinput.FileInput(file_name, inplace=True, backup='.bak') as file:
        for line in file:
            print(line.replace("117972993", lb_serial), end='')
        file.close()
    test_log.info("Ladybird Serial is identified and Changed in the "
                  "config.xml --------------------------------------- [OK]")


def installation_python():

    python_avail_res = python_check()

    if not python_avail_res:
        test_log.info("Python not installed, cross checking Python installation again...")

    uninstall_python_host()

    python_res = install_python()

    install_pip()

    if python_res:
        test_log.info("Python installation done, please go to the main page and proceed with git installation..")
    else:
        test_log.error("Python installation failed, please contact the developer for debugging - Naveenraj,K..")

    test_log.info('\n+==============================================================================================+\n'
                  '+==============================================================================================+')

    install_python_button.destroy()


def powercli_installation():

    test_log.info("PowerCLI installation started..")
    try:
        sys_path = Path(os.path.join(__file__)).parent
        py_file_name = os.path.join(os.path.join(sys_path, "lib"), "powercli.ps1")

        p = subprocess.Popen(["powershell", "-c", py_file_name], stdout=sys.stdout)
        p.communicate()

        if p.returncode == 0:
            test_log.info("PowerCLI installation successfull..")
            return True
        else:
            test_log.info("PowerCLI installation Failed..")
            return False
    except Exception as ex:
        test_log.error(ex)


def flask_server_start():
    try:
        test_log.info("Starting the Flask service..")
        python_path = "C:\Python38"
        # if str(sys_config_dict["cpu_stepping"]).startswith("E2"):
        #     python_path = "C:\Python38"
        # else:
        #     python_path = "C:\Python36"

        sys_path = Path(os.path.join(__file__)).parent
        py_file_name = os.path.join(os.path.join(sys_path, "lib"), "flask_server_flex.py")

        subprocess.Popen('{}\python.exe {}'.format(python_path, py_file_name),
                         creationflags=subprocess.CREATE_NEW_CONSOLE)
        test_log.info("Flask service started..")
    except Exception as ex:
        test_log.error(ex)


def nuc_provisioning():

    root.destroy()

    platform_chosen = get_platform_type()

    cpld_application_path = quartus_checkup()

    if not any(is_var_empty):
        generate_system_config_xml(platform_chosen, cpld_application_path)

    test_log.info('+==============================================================================================+')
    # add_reg_sz_key_env_path("GIT_PYTHON_GIT_EXECUTABLE", "C:\\Program Files\\Git\\bin\\git.exe")

    # Navee - Done
    time.sleep(2)
    cc_build_download()
    # pull_git_repos()

    test_log.info('+==============================================================================================+')

    # Navee - Done
    time.sleep(2)
    copy_config_from_dtaf_content_framework("content_configuration.xml")
    time.sleep(2)

    test_log.info('+==============================================================================================+')

    copy_config_from_dtaf_content_framework("content_configuration_seamless.xml")
    time.sleep(2)

    test_log.info('+==============================================================================================+')

    copy_config_from_dtaf_content_framework("miv_config.cfg")
    time.sleep(2)

    test_log.info('+==============================================================================================+')

    # installing python modules
    time.sleep(2)
    install_python_modules()

    test_log.info('+==============================================================================================+')

    dtaf_content_req_path = 'C:\\DPG_Automation\\dtaf_content\\requirements_py3.txt'
    time.sleep(2)
    dtaf_core_req_path = "C:\\DPG_Automation\\dtaf_core\\requirements\\requirements_ut_py37.txt"

    # python modules check
    time.sleep(2)
    if os.path.isfile(dtaf_content_req_path):
        python_lib_check(dtaf_content_req_path)
    else:
        test_log.info("Requirements Path not found: '{}'".format(dtaf_content_req_path))

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    if os.path.isfile(dtaf_core_req_path):
        python_lib_check(dtaf_core_req_path)
    else:
        test_log.info("Requirements Path not found: '{}'".format(dtaf_core_req_path))

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    cscript_installer()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    pythonsv_installer_semi_auto()

    test_log.info('+==============================================================================================+')

    # Navee - Done
    time.sleep(2)
    verify_com_port(sys_config_dict["bios_com_port_number"])
    time.sleep(2)
    verify_com_port(sys_config_dict["bmc_com_port_number"])

    test_log.info('+==============================================================================================+')

    # Navee - Done
    time.sleep(2)
    path_values = ["C:\DPG_Automation\dtaf_core\src", "C:\DPG_Automation\dtaf_content", "C:\DPG_Automation\montana",
                   "C:\pythonsv", "C:\cscripts"]
    pathname = "PYTHONPATH"
    add_reg_sz_key_env_path(pathname, path_values)

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    powercli_installation()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    # Navee - Done
    sprinter_res = install_sprinter_atf()

    if not sprinter_res:
        test_log.info("retrying one more time..")
        install_sprinter_atf()

    test_log.info('+==============================================================================================+')

    # Navee - Done
    time.sleep(2)
    install_notepad_res = install_notepad_plus_plus()

    if not install_notepad_res:
        test_log.info("retrying one more time..")
        install_notepad_plus_plus()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    itp_probe_check_device_manager()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    pycharm_installer()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    Putty_installer()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    winscp()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    banino_sx_sft_prepare_res = banino_sx_sft_prepare()

    if banino_sx_sft_prepare_res:
        banino_serial_setup()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    redfish_checkup()

    # test_log.info('+==============================================================================================+')
    #
    # time.sleep(2)
    # quartus_checkup()

    test_log.info('+==============================================================================================+')

    # pdu object created
    time.sleep(2)
    try:
        raritan_obj = PduDriver(test_log)
        raritan_obj.get_ac_power_state(timeout=20)
    except Exception as ex:
        test_log.error("Failed checking PDU working due to : {}".format(ex))

    test_log.info('+==============================================================================================+')
    
    flask_server_start()

    test_log.info('+==============================================================================================+')
    
    test_log.info("Nuc Provisioning Done..!!")

    test_log.info("SUT Provisioning Starting...")

    # Navee - Done - Only for windows SUT
    time.sleep(2)
    if sys_config_dict["sut_os_name"].lower() == "windows":
        test_log.info("SUT Type is Windows need to copy CCB Packages")
        ccb_res = install_ccb_package_xmlcli_windows_sut()
        if not ccb_res:
            test_log.info("retrying one more time..")
            install_ccb_package_xmlcli_windows_sut()
    else:
        test_log.info("Linux does not need CCB package on SUT.")

    test_log.info('+==============================================================================================+')

    # TODO
    # sut_copy_xmlcli_lin()

    time.sleep(2)
    sut_xmlcli_path_check()

    test_log.info('+==============================================================================================+')

    time.sleep(2)
    sut_python_check()

    test_log.info("SUT Provisioning Done!!!")

    test_log.info('+==============================================================================================+')


def get_verdict():
    """
    This method is used to parse the Auto installer log file to summarize the erros from the log.
    """
    try:
        parser_data_dict = {}
        with io.open(log_file_name, 'r') as auto_installer_log_file:
            data = auto_installer_log_file.read()

        error_message_parser = re.findall(r'.*ERROR\s+\[.*\](.*)', data, re.MULTILINE)
        test_log.info('+-----------------------------------------+')
        test_log.info('+           Errors & Warnings             +')
        test_log.info('+-----------------------------------------+')
        for error in error_message_parser:
            test_log.info("ERROR : {}".format(error))

        if len(error_message_parser) == 0:
            test_log.info("No errors Reported..")
            test_log.info("RESULT :: All Installation successfully Completed !!!!!!!!!")
        else:
            test_log.info("RESULT :: Some or all Installation failed to Complete !!!!!!!!!..., Please check the log "
                          "under --> '{}'".format(log_file_name))
    except Exception as ex:
        test_log.debug(ex)

def disable_quick_edit_mode():
    import platform

    if platform.system() != "windows":
        return
    self._log.info("Windows OS: Disable command prompt quick edit mode.")
    import ctypes
    from ctypes import wintypes
    kernel32 = ctypes.WinDLL('kernel32')
    dword_for_std_input_handle = ctypes.wintypes.DWORD(-10)
    dword_for_enable_extended_flags = ctypes.wintypes.DWORD(0x0080)
    std_input_handle = kernel32.GetStdHandle(dword_for_std_input_handle)
    kernel32.SetConsoleMode(std_input_handle, dword_for_enable_extended_flags)
    le = kernel32.GetLastError()
    self._log.info("Windows OS: Disable command prompt quick edit mode completed with ret code:{}.".format(le))

def press_exit():
    try:
        from msvcrt import getch

        print("Press any key to exit...")
        getch()
    except Exception as ex:
        test_log.error(ex)


# Driver code
if __name__ == "__main__":
    try:

        # create a GUI window
        root = Tk()

        root.title("Host-SUT Preparation")

        # set the background colour of GUI window
        root.configure(background='light blue')

        # create a Form label
        heading = Label(root, text="*** Please select one platform among the below ***", bg='Green', fg='white')
        heading.grid(row=2, column=0)

        root.geometry("350x300")

        # create a 2 socket Button and place into the root window
        two_four_socket_button = Button(root, text="Generate 2S/4S System Configuration", padx=50,
                                        command=lambda button_pressed="2S_4S": open_platform_based_window(
                                            button_pressed))
        two_four_socket_button.grid(row=4, column=0, padx=5, pady=5)

        # create a 8 socket Button and place into the root window
        eight_socket_button = Button(root, text="Generate 8S System Configuration", padx=50,
                                     command=lambda button_pressed="8S": open_platform_based_window(button_pressed))
        eight_socket_button.grid(row=6, column=0, padx=5, pady=5)

        # create a 2 socket Button and place into the root window
        capi_two_four_socket_button = Button(root, text="(TBD) Generate CAPI System Configuration", padx=50,
                                        command=lambda button_pressed="2S_4S": open_platform_based_window(
                                            button_pressed), bg="yellow")
        capi_two_four_socket_button.grid(row=8, column=0, padx=5, pady=5)

        # create a Form label
        heading = Label(root, text="*** Once System Configuration is done, \n "
                                   "please start below setups one after another ***", bg='Green', fg='white')
        heading.grid(row=10, column=0)

        # create a python installation Button and place into the root window
        install_python_button = Button(root, text="Install Python", padx=50, command=installation_python)
        install_python_button.grid(row=12, column=0, padx=5, pady=5)

        # create a git installation Button and place into the root window
        install_git_button = Button(root, text="Install Git-Bash", padx=50, command=git_bash)
        install_git_button.grid(row=14, column=0, padx=5, pady=5)

        # create a Nuc setup installation Button and place into the root window
        nuc_setup = Button(root, text="NUC Setup", padx=50, command=nuc_provisioning)
        nuc_setup.grid(row=16, column=0, padx=5, pady=5)

        # start the GUI
        root.mainloop()
    except Exception as ex:
        test_log.error(ex)
    finally:
        get_verdict()
        installer_info()
        press_exit()
        # File Explorer will open the path where log file is stored.
        subprocess.Popen(r'explorer /select,{}'.format(log_file_name))
